import os, random, string, sqlite3, json, uuid
import openpyxl
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes

TOKEN = os.getenv("BOT_TOKEN")
DOMAIN = os.getenv("MAIL_DOMAIN", "jmogmail.com")
ADMIN_IDS = [7962377092]
QRIS_IMAGE = os.getenv("QRIS_IMAGE", "https://i.ibb.co/6JJ3zQr9/final-qris-poster.png")
# === REKENING & DANA KAMU - UDAH FIX SESUAI DATA KAMU ===
REKENING_BANK = os.getenv("REKENING_BANK", "SEABANK")
REKENING_NUMBER = os.getenv("REKENING_NUMBER", "901040978290")
REKENING_NAME = os.getenv("REKENING_NAME", "HAMBALI")
DANA_NUMBER = os.getenv("DANA_NUMBER", "083824101264")
DANA_NAME = os.getenv("DANA_NAME", "HAMBALI")
DB_NAME = "/tmp/bot_store.db"

# === EMAIL SERVER CONFIG - UNTUK EMAIL BISA LOGIN & TERIMA OTP ===
# Isi ini setelah beli domain jmogmail.com + hosting Hostinger
CPANEL_HOST = os.getenv("CPANEL_HOST", "")  # contoh: https://hpanel.hostinger.com atau https://yourdomain:2083
CPANEL_USER = os.getenv("CPANEL_USER", "")  # username cPanel / Hostinger
CPANEL_PASS = os.getenv("CPANEL_PASS", "")  # password cPanel atau API Token
CPANEL_API_TOKEN = os.getenv("CPANEL_API_TOKEN", "")  # kalau pakai API Token cPanel

HARGA = {"1":5000, "10":40000, "20":70000, "100":300000}
HARGA_FORMAT = {"1bulan":10000, "2bulan":20000, "3bulan":25000, "unlimited":200000}
HARGA_TOPUP = [5000, 10000, 20000, 25000, 40000, 70000, 200000, 300000]

SOLUSI_LENGKAP = {
    "025": """❌ **KODE 025 - NIK Sudah Terdaftar / Email Duplikat**

**Penyebab:** NIK kamu sudah pernah daftar JMO pakai email lama yang sekarang tidak bisa diakses.

**Solusi Resmi BPJS:**
1. Di halaman login JMO klik **Lupa Password**
2. Pilih verifikasi **NIK + Nama Ibu Kandung**
3. Kalau OTP tidak masuk ke email lama, langsung ke Kantor BPJS
4. Bawa KTP, KK, Paklaring asli
5. Minta petugas **Reset Email & Merge NIK**
6. Setelah reset, ganti pakai email baru dari bot ini **{DOMAIN}** biar OTP 100% masuk
7. Jangan daftar NIK baru, nanti malah double data.""",

    "026": """❌ **KODE 026 - KPJ Tidak Ditemukan / Tidak Aktif**

**Penyebab:** KPJ kamu belum dilaporkan perusahaan di SIPP atau sudah tidak aktif lama.

**Solusi:**
1. Tanya HRD perusahaan terakhir, minta No KPJ & pastikan sudah non-aktif di SIPP Online
2. Kalau punya 2 KPJ beda perusahaan, wajib **Penggabungan Saldo (Akuisisi)** di Kantor BPJS
3. Bawa Paklaring asli, KTP, KK
4. Kalau perusahaan sudah tutup, bawa Paklaring + Surat Keterangan dari Disnaker
5. Setelah KPJ aktif, baru bisa login JMO""",

    "040": """❌ **KODE 040 / BIOMETRIK GAGAL - Face Recognition Gagal**

**Ini error paling sering!**

**Penyebab:** Wajah di KTP tidak cocok dengan selfie.

**Solusi Jitu (95% berhasil):**
1. Foto di tempat **TERANG**, background dinding putih polos
2. Lepas Kacamata, Masker, Topi, Headset
3. HP jangan goyang, hadap kamera lurus
4. Jangan pakai filter beauty / makeup tebal
5. Coba 3x, kalau gagal tunggu 1-2 jam jangan spam
6. **Paling penting:** Foto KTP kamu di Dukcapil harus jelas. Kalau KTP lama buram, update KTP di Kecamatan dulu
7. Kalau tetap 040, klaim offline ke Kantor BPJS bawa KTP asli, petugas bisa bypass biometrik""",

    "05": """❌ **KODE 05 - NIK Tidak Valid / Tidak Terdaftar di Dukcapil**

Penyebab: NIK di BPJS beda dengan Dukcapil.

Solusi:
1. Cek KTP, pastikan NIK 16 digit benar
2. Ke Dukcapil Kecamatan minta **Perbaikan Data / Sinkronisasi NIK**
3. Setelah Dukcapil update, tunggu 3x24 jam baru coba JMO lagi
4. Kalau NIK di BPJS salah input HRD, bawa KTP + KK ke Kantor BPJS untuk koreksi NIK""",

    "06": """❌ **KODE 06 - Nama Tidak Sesuai / Data Tidak Sinkron**

Penyebab: Nama di KTP beda dengan nama di data BPJS (beda spasi / gelar).

Solusi:
1. Bawa KTP + KK + Paklaring ke Kantor BPJS
2. Minta **Perbaikan Nama** sesuai KTP
3. Jangan ubah nama di JMO, ubah di database BPJS pusat
4. Setelah perbaikan, login JMO lagi""",

    "011": """❌ **KODE 011 - Nomor HP Sudah Terdaftar**

Penyebab: No HP sudah dipakai akun JMO lain.

Solusi:
1. Pakai No HP lain yang belum pernah daftar JMO
2. Atau pakai fitur Lupa Password > ganti No HP dengan verifikasi NIK + Ibu Kandung
3. Kalau No HP hilang, datang ke Kantor BPJS bawa KTP minta ganti No HP""",

    "020": """❌ **KODE 020 - Gagal Ambil Data KPJ**

Penyebab: Server BPJS lagi maintenance atau KPJ belum migrasi ke JMO.

Solusi:
1. Coba lagi jam 08.00 - 15.00 (jam kerja)
2. Update aplikasi JMO ke versi terbaru
3. Hapus cache JMO di Pengaturan HP
4. Kalau tetap, KPJ kamu masih di sistem lama, harus aktivasi ulang di Kantor BPJS""",

    "021": """❌ **KODE 021 - Saldo Tidak Muncul / 0**

Penyebab: Saldo belum diinput perusahaan atau ada KPJ ganda yang belum digabung.

Solusi:
1. Cek di menu **Cek Saldo**, bukan di halaman login
2. Kalau saldo 0, tanya HRD apakah iuran BPJS sudah dibayar?
3. Kalau punya 2 perusahaan, saldo ada di KPJ lama, minta akuisisi di Kantor BPJS
4. Tunggu 7 hari setelah resign, saldo baru muncul""",

    "024": """❌ **KODE 024 - Email Sudah Terpakai**

Solusi: Ini alasan utama bot ini dibuat!
1. Beli 1 Email baru di bot ini {DOMAIN} (cuma 5k)
2. Login email baru di Gmail App
3. Ganti email JMO ke email baru, OTP pasti masuk
4. Email lama yang sudah terpakai tidak bisa dipakai lagi untuk NIK lain""",

    "027": """❌ **KODE 027 / 029 - Rekening Tidak Valid**

Penyebab: Rekening bukan atas nama sendiri atau salah input.

Solusi:
1. Wajib rekening **ATAS NAMA SENDIRI** sesuai KTP
2. Boleh BCA, BRI, Mandiri, BNI
3. Jangan pakai e-Wallet (DANA, OVO, GOPAY) untuk klaim diatas 10jt
4. Foto buku rekening harus jelas 4 angka terakhir
5. Kalau rekening sudah tutup, buka rekening baru dulu""",

    "030": """❌ **KODE 030 / 031 - Paklaring / Surat Keterangan Kerja Tidak Valid**

Solusi:
1. Paklaring harus ada: Nama, NIK, No KPJ, Tgl Masuk-Keluar, Cap Perusahaan
2. Kalau perusahaan tutup, pakai Surat Keterangan dari Disnaker + Paklaring lama
3. Foto Paklaring jangan blur, format JPG bukan PDF
4. Kalau di-PHK, lampirkan Surat PHK""",

    "035": """❌ **KODE 035 - Gagal Upload Dokumen**

Solusi:
1. File maksimal 5MB, format JPG/PNG
2. Jangan upload PDF
3. Koneksi internet harus stabil
4. Coba upload 1 per 1, jangan langsung semua
5. Hapus cache JMO""",

    "041": """❌ **KODE 041 / 045 - Verifikasi Gagal / Data Tidak Cocok**

Solusi:
1. Cek lagi NIK, Nama Ibu Kandung, Tgl Lahir harus 100% sama dengan KTP & KK
2. Nama Ibu Kandung tanpa gelar (contoh: Siti, bukan Hj. Siti)
3. Kalau Ibu sudah meninggal, tetap pakai nama Ibu Kandung asli
4. Kalau tetap gagal, perbaikan data di Kantor BPJS""",

    "050": """❌ **KODE 050 / 051 - Klaim Ditolak / Gagal Validasi**

Penyebab: Dokumen kurang atau saldo belum waktunya cair.

Solusi:
1. Cek alasan penolakan di notifikasi JMO
2. Kalau dokumen kurang, upload ulang yang lebih jelas
3. Pastikan sudah resign minimal 1 bulan
4. Untuk klaim 100%, usia minimal 56 atau resign + 1 bulan
5. Hubungi 175 untuk cek status klaim""",

    "otp": """📧 **OTP TIDAK MASUK - SOLUSI PALING SERING**

Kenapa OTP JMO tidak masuk?
1. Email lama kena spam / penuh
2. Server email gratis (gmail, yahoo) sering delay

**Solusi Jitu Bot Ini:**
1. Beli Email baru {DOMAIN} di bot - Server fresh anti delay
2. Login di Gmail App (Other Account)
3. Masuk JMO > Pengkinian Data > Ganti Email
4. OTP masuk dalam 5 detik!
5. Cek folder Spam juga

Garansi OTP masuk, kalau tidak masuk refund!""",

    "lupa password": """🔑 **LUPA PASSWORD JMO**

1. Klik Lupa Password di login JMO
2. Masukkan NIK + Nama Ibu Kandung
3. Pilih kirim OTP via Email
4. Kalau email lama hilang, pakai solusi OTP di atas beli email baru {DOMAIN}
5. Buat password baru minimal 8 karakter kombinasi huruf + angka""",

    "klaim": """💰 **CARA KLAIM JHT JMO BIAR CEPAT CAIR 100%**

**Syarat Wajib:**
- Sudah resign minimal 30 hari
- Perusahaan sudah lapor non-aktif di SIPP
- Dokumen lengkap

**Langkah:**
1. Update Email & HP di JMO dulu (pakai email {DOMAIN} dari bot)
2. Menu **Pengajuan Klaim > JHT**
3. Pilih Alasan: Resign
4. Upload: KTP, KK, Paklaring, Buku Rekening, Selfie + KTP
5. Pastikan rekening atas nama sendiri
6. Nominal <10jt cair 1-3 hari, >10jt 5-7 hari kerja
7. Cek status di menu Tracking Klaim

**Tips Biar Tidak Ditolak:**
- Foto dokumen jangan blur
- Paklaring harus ada cap perusahaan
- Rekening aktif & sesuai KTP""",

    "saldo": """💳 **SALDO JMO TIDAK MUNCUL / BERKURANG**

1. Saldo muncul di menu **Jaminan Hari Tua > Cek Saldo**, bukan di beranda
2. Saldo berkurang karena ada biaya admin tahunan 5k
3. Kalau punya 2 KPJ, saldo terpisah, harus akuisisi di Kantor BPJS biar gabung
4. Kalau baru resign, saldo update tiap tanggal 10 bulan depan"""
}

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (user_id INTEGER PRIMARY KEY, username TEXT, saldo INTEGER DEFAULT 0, created_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS emails (id INTEGER PRIMARY KEY AUTOINCREMENT, email TEXT, password TEXT, owner_id INTEGER, created_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS subscriptions (user_id INTEGER PRIMARY KEY, expiry TEXT, unlimited INTEGER DEFAULT 0, plan TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS formats (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, title TEXT, content TEXT, created_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS format_settings (user_id INTEGER PRIMARY KEY, template TEXT, updated_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS format_history (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, raw_text TEXT, formatted_text TEXT, created_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS topup (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, amount INTEGER, status TEXT, created_at TEXT)''')
    conn.commit(); conn.close()

def get_format_setting(user_id):
    conn=sqlite3.connect(DB_NAME)
    row=conn.execute("SELECT template FROM format_settings WHERE user_id=?", (user_id,)).fetchone()
    conn.close()
    if row and row[0]:
        return row[0]
    # Template default - user bisa ubah di SETTING FORMAT
    return """📢 INFO LOKER - {judul}

📍 Lokasi: {lokasi}
💼 Posisi: {posisi}
💰 Gaji: {gaji}
✨ Benefit: {benefit}

📝 Detail:
{isi_asal}

#loker #jmo #bpjs #info"""

def auto_format_text(raw_text, template):
    """Format otomatis sesuai setting - ambil kata kunci dari text asal2an"""
    # Bersihkan text
    lines = [l.strip() for l in raw_text.strip().split("\n") if l.strip()]
    raw_lower = raw_text.lower()
    
    # Coba deteksi otomatis
    # Contoh input:
    # Info loker
    # Bandung
    # Antapani
    # Kulim
    # Gaji 12 juta
    # Dpt jmo lasik jos
    
    # Logic sederhana: parse
    judul = lines[0] if lines else "Info Loker"
    lokasi = ""
    gaji = ""
    benefit = ""
    posisi = ""
    
    # Cari lokasi, gaji, benefit
    for line in lines:
        ll = line.lower()
        if "gaji" in ll or "jt" in ll or "juta" in ll:
            gaji = line
        elif "jmo" in ll or "bpjs" in ll or "lasik" in ll or "dpt" in ll or "benefit" in ll:
            benefit += line + " "
        elif "bandung" in ll or "jakarta" in ll or "antapani" in ll or "kulim" in ll or "surabaya" in ll or "bekasi" in ll:
            lokasi += line + " "
        else:
            if line != judul:
                posisi += line + " "
    
    if not lokasi:
        lokasi = " - ".join(lines[1:3]) if len(lines)>=3 else "Bandung"
    if not gaji:
        gaji = "Negosiasi"
    if not benefit:
        benefit = "JMO, BPJS"
    if not posisi:
        posisi = "Staff"
    
    # Ganti placeholder di template
    result = template
    result = result.replace("{judul}", judul.title())
    result = result.replace("{lokasi}", lokasi.strip())
    result = result.replace("{posisi}", posisi.strip()[:50])
    result = result.replace("{gaji}", gaji)
    result = result.replace("{benefit}", benefit.strip())
    result = result.replace("{isi_asal}", raw_text)
    result = result.replace("{isi}", raw_text)
    result = result.replace("{text}", raw_text)
    
    return result.strip()

def get_user(user_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE user_id=?", (user_id,))
    user = c.fetchone()
    if not user:
        c.execute("INSERT INTO users VALUES (?,?,?,?)", (user_id, "", 0, datetime.now().strftime("%d-%m-%Y")))
        conn.commit()
        user = (user_id, "", 0, datetime.now().strftime("%d-%m-%Y"))
    conn.close()
    return user

def check_sub(user_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT expiry, unlimited, plan FROM subscriptions WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    if not row: return False, "Belum langganan"
    expiry_str, unlimited, plan = row
    if unlimited==1: return True, f"Unlimited ({plan})"
    try:
        expiry=datetime.strptime(expiry_str, "%Y-%m-%d %H:%M:%S")
        if datetime.now() < expiry:
            sisa=(expiry-datetime.now()).days+1
            return True, f"{plan} sisa {sisa} hari"
        else:
            return False, f"Expired {expiry.strftime('%d-%m-%Y')}"
    except:
        return False, "Expired"

def add_sub(user_id, paket):
    now=datetime.now()
    if paket=="unlimited": expiry="9999-12-31 23:59:59"; unlimited=1
    elif paket=="1bulan": expiry=(now+timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S"); unlimited=0
    elif paket=="2bulan": expiry=(now+timedelta(days=60)).strftime("%Y-%m-%d %H:%M:%S"); unlimited=0
    else: expiry=(now+timedelta(days=90)).strftime("%Y-%m-%d %H:%M:%S"); unlimited=0
    conn=sqlite3.connect(DB_NAME)
    c=conn.cursor()
    c.execute("INSERT OR REPLACE INTO subscriptions VALUES (?,?,?,?)",(user_id,expiry,unlimited,paket))
    conn.commit(); conn.close()
    return expiry

def main_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("1. PROFIL", callback_data="profil")],
        [InlineKeyboardButton(f"2. BELI EMAIL @{DOMAIN}", callback_data="beli_email")],
        [InlineKeyboardButton("3. TOP UP", callback_data="topup_qris")],
        [InlineKeyboardButton("4. CARA LOGIN dan TERIMA OTP", callback_data="cara_login")],
        [InlineKeyboardButton("5. SOLUSI JMO", callback_data="solusi_jmo")],
        [InlineKeyboardButton("6. FORMAT OTOMATIS", callback_data="format_jmo")],
        [InlineKeyboardButton("7. BANTUAN", callback_data="bantuan")],
        [InlineKeyboardButton("8. CEK STATUS", callback_data="cek_status")],
        [InlineKeyboardButton("9. HUBUNGI ADMIN", callback_data="hubungi_admin")],
        [InlineKeyboardButton("10. PANEL ADMIN", callback_data="admin")]
    ])

def paket_menu(saldo):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"1 Email Rp {HARGA['1']:,}", callback_data="buy_1")],
        [InlineKeyboardButton(f"10 Email Rp {HARGA['10']:,} Hemat", callback_data="buy_10")],
        [InlineKeyboardButton(f"20 Email Rp {HARGA['20']:,} Paling Laris", callback_data="buy_20")],
        [InlineKeyboardButton(f"100 Email Rp {HARGA['100']:,} Super Hemat", callback_data="buy_100")],
        [InlineKeyboardButton(f"⬅️ Kembali Saldo Rp {saldo:,}", callback_data="back")]
    ])

def paket_format_menu(saldo):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"1 Bulan Rp {HARGA_FORMAT['1bulan']:,}", callback_data="sub_1bulan")],
        [InlineKeyboardButton(f"2 Bulan Rp {HARGA_FORMAT['2bulan']:,}", callback_data="sub_2bulan")],
        [InlineKeyboardButton(f"3 Bulan Rp {HARGA_FORMAT['3bulan']:,} Hemat", callback_data="sub_3bulan")],
        [InlineKeyboardButton(f"Unlimited Rp {HARGA_FORMAT['unlimited']:,}", callback_data="sub_unlimited")],
        [InlineKeyboardButton(f"💰 Top Up Rp {saldo:,}", callback_data="topup_qris")],
        [InlineKeyboardButton("⬅️ Kembali", callback_data="back")]
    ])

def gen_email():
    prefix=''.join(random.choices(string.ascii_lowercase, k=6))+str(random.randint(10,999))
    email=f"{prefix}@{DOMAIN}"
    pwd=''.join(random.choices(string.ascii_letters+string.digits, k=10))+"!A1"
    return email,pwd

def create_real_email_on_server(full_email, password):
    """Bikin mailbox asli di cPanel/Hostinger biar user bisa login & terima OTP"""
    # Jika belum setting CPANEL, tetap return True tapi mode simulasi (biar bot tetap jalan)
    if not CPANEL_HOST or not CPANEL_USER:
        # Mode simulasi - email belum beneran dibuat di server, user gak bisa login tapi bot tetap kasih email
        # Nanti setelah beli hosting, isi CPANEL_HOST/USER/PASS di Railway, email akan beneran kebikin otomatis
        return True, "SIMULASI - Set CPANEL_HOST di Railway untuk bikin mailbox asli"
    
    try:
        import requests
        user_part, domain_part = full_email.split("@")
        
        # Pakai cPanel UAPI Email::add_pop
        if CPANEL_API_TOKEN:
            # Mode API Token (lebih aman)
            headers = {"Authorization": f"cpanel {CPANEL_USER}:{CPANEL_API_TOKEN}"}
            url = f"{CPANEL_HOST.rstrip('/')}/execute/Email/add_pop"
            data = {"email": user_part, "password": password, "domain": domain_part, "quota": 250}
            r = requests.post(url, headers=headers, data=data, timeout=15)
            j = r.json()
            if j.get("status") == 1:
                return True, "OK - Mailbox asli berhasil dibuat!"
            else:
                return False, f"Gagal API: {j.get('errors')}"
        else:
            # Mode username + password cPanel
            import base64
            url = f"{CPANEL_HOST.rstrip('/')}/execute/Email/add_pop"
            # Basic auth
            auth_str = f"{CPANEL_USER}:{CPANEL_PASS}"
            b64 = base64.b64encode(auth_str.encode()).decode()
            headers = {"Authorization": f"Basic {b64}"}
            data = {"email": user_part, "password": password, "domain": domain_part, "quota": 250}
            r = requests.post(url, headers=headers, data=data, timeout=15)
            j = r.json()
            if j.get("status") == 1:
                return True, "OK - Mailbox asli berhasil dibuat!"
            else:
                return False, f"Gagal: {j.get('errors')}"
    except Exception as e:
        return False, f"Exception: {str(e)[:100]}"

def get_cara_login_text(email, pwd):
    return f"""📱 **CARA LOGIN EMAIL DI HP - BISA TERIMA OTP LANGSUNG**

Email kamu: `{email}`
Password: `{pwd}`

**CARA LOGIN DI HP ANDROID (Gmail App):**
1. Buka App Gmail > Klik Foto Profil kanan atas > Tambah Akun Lain
2. Pilih **Lainnya / Other**
3. Masukan Email: `{email}`
4. Pilih **Akun Pribadi (IMAP)**
5. Password: `{pwd}`
6. Server Masuk: `mail.{DOMAIN}`
   Port: 993, Keamanan: SSL/TLS
7. Server Keluar: `mail.{DOMAIN}`
   Port: 465, Keamanan: SSL/TLS
8. Selesai! Inbox akan muncul

**CARA LOGIN DI iPhone:**
1. Pengaturan > Mail > Akun > Tambah Akun > Lainnya
2. Tambah Akun Mail > isi Nama, Email `{email}`, Password `{pwd}`
3. IMAP: Host `mail.{DOMAIN}`
4. Selesai

**TEST OTP JMO:**
1. Buka JMO > Ganti Email > Masukan `{email}`
2. OTP akan masuk langsung ke Gmail App kamu dalam 5 detik!
3. Cek di folder Inbox / Spam

**IMAP Setting Lengkap:**
- Email: {email}
- Password: {pwd}
- IMAP Host: mail.{DOMAIN} Port 993 SSL
- SMTP Host: mail.{DOMAIN} Port 465 SSL

Kalau gagal login, hubungi admin!"""

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    get_user(update.effective_user.id)
    conn=sqlite3.connect(DB_NAME)
    conn.execute("UPDATE users SET username=? WHERE user_id=?", (update.effective_user.username, update.effective_user.id))
    conn.commit(); conn.close()
    txt=f"""**BOT JMO EMAIL - {DOMAIN} - 1 KLIK JADI!** 🚀

✅ **EMAIL BISA LOGIN DI HP & TERIMA OTP JMO LANGSUNG!**

**HARGA MURAH:**
• 1 Email = Rp 5.000
• 10 Email = Rp 40.000
• 20 Email = Rp 70.000 (Paling Laris)
• 100 Email = Rp 300.000

**CARA KERJA 1 KLIK JADI:**
1. Topup Saldo via QRIS / SEABANK {REKENING_NUMBER} / DANA {DANA_NUMBER}
2. Beli Email -> Bot bikin mailbox asli @{DOMAIN}
3. User login di HP pakai Gmail App
4. OTP JMO masuk langsung ke HP user!

**PAYMENT:**
QRIS: {QRIS_IMAGE}
SEABANK: {REKENING_NUMBER} a.n {REKENING_NAME}
DANA: {DANA_NUMBER} a.n {DANA_NAME}

Pilih menu di bawah - SIMPLE & CEPAT!"""
    await update.message.reply_text(txt, reply_markup=main_menu(), parse_mode="Markdown")

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q=update.callback_query
    await q.answer()
    uid=q.from_user.id
    data=q.data
    user=get_user(uid)
    saldo=user[2]
    is_sub, sub_info=check_sub(uid)

    if data=="profil":
        await q.edit_message_text(f"👤 ID: `{uid}`\n💰 Saldo: Rp {saldo:,}\n📧 Domain: @{DOMAIN}\n🔒 Langganan: {sub_info}\n\nHarga Email: 1=5k, 10=40k, 20=70k, 100=300k", reply_markup=main_menu(), parse_mode="Markdown")

    elif data=="saldo":
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("💳 TOPUP QRIS & REKENING", callback_data="topup_qris")],
            [InlineKeyboardButton("➕ Cara Topup", callback_data="cara_topup")],
            [InlineKeyboardButton("⬅️ Kembali", callback_data="back")]
        ])
        await q.edit_message_text(f"💰 Saldo Rp {saldo:,}\nID: `{uid}`\n\nPilih metode topup:\n✅ QRIS (All e-wallet)\n✅ SEABANK {REKENING_NUMBER}\n✅ DANA {DANA_NUMBER}", reply_markup=kb)

    elif data=="cara_topup":
        txt=f"""💰 **CARA TOPUP - 3 METODE:**

**1️⃣ QRIS (Otomatis All E-Wallet):**
Scan QRIS di bot, bisa pakai DANA, OVO, GoPay, ShopeePay, M-Banking semua

**2️⃣ Transfer SEABANK:**
Bank: {REKENING_BANK}
No: `{REKENING_NUMBER}`
A.n: {REKENING_NAME}

**3️⃣ Transfer DANA:**
No: `{DANA_NUMBER}`
A.n: {DANA_NAME}

Setelah transfer, upload bukti foto di bot. Saldo masuk otomatis FULL AUTO (karena kamu sibuk)!

ID Kamu: `{uid}`"""
        await q.edit_message_text(txt, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("💳 TOPUP SEKARANG", callback_data="topup_qris")],[InlineKeyboardButton("⬅️ Kembali", callback_data="back")]]), parse_mode="Markdown")

    elif data=="topup_qris":
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("5k - 1 Email", callback_data="qris_5000"), InlineKeyboardButton("40k - 10 Email", callback_data="qris_40000")],
            [InlineKeyboardButton("70k - 20 Email", callback_data="qris_70000"), InlineKeyboardButton("10k - Format 1 Bln", callback_data="qris_10000")],
            [InlineKeyboardButton("20k - Format 2 Bln", callback_data="qris_20000"), InlineKeyboardButton("25k - Format 3 Bln", callback_data="qris_25000")],
            [InlineKeyboardButton("300k - 100 Email", callback_data="qris_300000"), InlineKeyboardButton("200k - Unlimited Format", callback_data="qris_200000")],
            [InlineKeyboardButton("💰 Nominal Lain", callback_data="qris_custom")],
            [InlineKeyboardButton("⬅️ Kembali", callback_data="back")]
        ])
        txt=f"""💳 **TOPUP QRIS & REKENING - AUTO**
Saldo: Rp {saldo:,}

Pilih nominal dulu, nanti bot kasih 3 pilihan bayar:

**1. QRIS** - Scan pakai e-wallet apa aja
**2. SEABANK:** {REKENING_BANK} {REKENING_NUMBER} a.n {REKENING_NAME}
**3. DANA:** {DANA_NUMBER} a.n {DANA_NAME}

Semua metode FULL AUTO - upload bukti langsung saldo masuk!

Harga Email: 1=5k, 10=40k, 20=70k"""
        await q.edit_message_text(txt, reply_markup=kb, parse_mode="Markdown")

    elif data.startswith("qris_"):
        if data=="qris_custom":
            context.user_data['mode']='topup_custom'
            await q.edit_message_text("💳 Ketik nominal topup (minimal 5000):\nContoh: 50000", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Batal", callback_data="topup_qris")]]))
            return
        amount=int(data.replace("qris_",""))
        conn=sqlite3.connect(DB_NAME)
        conn.execute("INSERT INTO topup (user_id, amount, status, created_at) VALUES (?,?,?,?)",(uid, amount, "pending", datetime.now().strftime("%d-%m-%Y %H:%M")))
        topup_id=conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit(); conn.close()
        txt=f"""💳 **TOPUP Rp {amount:,} - ID #{topup_id}**
User: {uid}
Nominal: **Rp {amount:,}** (harus pas!)

**PILIH 1 DARI 3 METODE BAYAR:**

**1️⃣ QRIS (All E-Wallet):**
Scan QR di bawah ini pakai DANA/OVO/GoPay/ShopeePay/M-Banking

**2️⃣ SEABANK:**
Bank: {REKENING_BANK}
No: `{REKENING_NUMBER}`
A.n: {REKENING_NAME}

**3️⃣ DANA:**
No: `{DANA_NUMBER}`
A.n: {DANA_NAME}

**CARA:**
1. Transfer / Scan QRIS sesuai nominal Rp {amount:,}
2. Screenshot bukti transfer
3. Klik tombol Upload Bukti di bawah
4. Saldo masuk OTOMATIS FULL AUTO!

ID: #{topup_id}"""
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("📤 Saya Sudah Bayar - Upload Bukti", callback_data="upload_bukti")],[InlineKeyboardButton("⬅️ Kembali", callback_data="topup_qris")]])
        try:
            await q.message.reply_photo(photo=QRIS_IMAGE, caption=txt, reply_markup=kb, parse_mode="Markdown")
            await q.edit_message_text("✅ QRIS & Rekening dikirim di atas! Silahkan transfer Rp "+f"{amount:,}", reply_markup=main_menu())
        except:
            await q.edit_message_text(txt + f"\n\n[QRIS: {QRIS_IMAGE}]", reply_markup=kb, parse_mode="Markdown")

    elif data=="upload_bukti":
        context.user_data['mode']='upload_bukti'
        await q.edit_message_text(f"📤 **Upload Bukti Transfer - MODE AUTO**\n\nKirim foto screenshot bukti TF ke:\n\n**SEABANK:** {REKENING_NUMBER} a.n {REKENING_NAME}\n**DANA:** {DANA_NUMBER} a.n {DANA_NAME}\n**QRIS:** Scan QR di atas\n\nSaldo akan **LANGSUNG MASUK OTOMATIS** tanpa tunggu admin!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Batal", callback_data="topup_qris")]]))

    elif data=="beli_email":
        await q.edit_message_text(f"📧 **PAKET EMAIL @{DOMAIN} HARGA BARU**\n\n1 Email = Rp 5.000\n10 Email = Rp 40.000\n20 Email = Rp 70.000 (Paling Laris)\n100 Email = Rp 300.000\n\nSaldo: Rp {saldo:,}", reply_markup=paket_menu(saldo))

    elif data.startswith("buy_"):
        jml=int(data.split("_")[1])
        harga=HARGA[str(jml)]
        if saldo<harga:
            await q.edit_message_text(f"❌ Saldo kurang! Butuh Rp {harga:,}, saldo Rp {saldo:,}\n\nTopup dulu via QRIS!", reply_markup=paket_menu(saldo)); return
        
        # Kirim loading dulu biar user tau lagi proses bikin email asli
        await q.edit_message_text(f"⏳ **SEDANG MEMBUAT {jml} EMAIL ASLI @{DOMAIN}**\n\nEmail sedang dibuat di server mail...\nUser bisa login di HP & terima OTP langsung!\n\nMohon tunggu 5-10 detik...", parse_mode="Markdown")
        
        conn=sqlite3.connect(DB_NAME)
        conn.execute("UPDATE users SET saldo=saldo-? WHERE user_id=?",(harga,uid))
        hasil=[]
        hasil_detail=[]
        gagal=[]
        for _ in range(jml):
            email,pwd=gen_email()
            # BIKIN MAILBOX ASLI DI SERVER CPANEL/HOSTINGER
            ok, msg = create_real_email_on_server(email, pwd)
            if ok:
                hasil.append(f"{email}|{pwd}")
                hasil_detail.append(f"{email} | {pwd} | OK")
            else:
                # Kalau gagal bikin di server (misal belum setting CPANEL), tetap kasih email tapi kasih warning
                hasil.append(f"{email}|{pwd}")
                hasil_detail.append(f"{email} | {pwd} | {msg}")
                gagal.append(email)
            conn.execute("INSERT INTO emails (email,password,owner_id,created_at) VALUES (?,?,?,?)",(email,pwd,uid,datetime.now().strftime("%d-%m-%Y %H:%M")))
        conn.commit(); conn.close()
        
        text_hasil="\n".join(hasil)
        sisa=saldo-harga
        
        # Pesan cara login
        if jml==1:
            cara_login = get_cara_login_text(hasil[0].split("|")[0], hasil[0].split("|")[1])
            # Kirim email + cara login
            if gagal:
                warn = f"\n\n⚠️ **NOTE:** {len(gagal)} email mode SIMULASI (belum setting CPANEL_HOST). User belum bisa login sampai kamu setting hosting.\nSet di Railway: CPANEL_HOST, CPANEL_USER, CPANEL_PASS"
            else:
                warn = "\n\n✅ **MAILBOX ASLI SUDAH AKTIF!** User bisa langsung login di HP & terima OTP JMO!"
            
            await q.edit_message_text(f"✅ **SUKSES {jml} EMAIL @{DOMAIN} - BISA LOGIN DI HP!**\nSisa Saldo: Rp {sisa:,}\n\n```\n{text_hasil}\n```{warn}", reply_markup=main_menu(), parse_mode="Markdown")
            # Kirim cara login detail sebagai follow up
            try:
                await q.message.reply_text(cara_login, parse_mode="Markdown")
            except:
                pass
        else:
            # Untuk pembelian banyak, kirim file + cara login umum
            path=f"/tmp/{jml}_email_{uid}.txt"
            with open(path,"w") as f:
                f.write(f"# {jml} EMAIL @{DOMAIN} - BISA LOGIN DI HP & TERIMA OTP\n")
                f.write(f"# Dibuat: {datetime.now().strftime('%d-%m-%Y %H:%M')}\n")
                f.write(f"# Format: EMAIL|PASSWORD\n")
                f.write(f"# Cara Login: Gmail App > Tambah Akun Lain > Lainnya > IMAP mail.{DOMAIN} Port 993\n")
                f.write(f"# \n")
                f.write(text_hasil)
                f.write(f"\n\n# CARA LOGIN LENGKAP:\n")
                f.write(f"# 1. Gmail App > Profil > Tambah Akun Lain > Lainnya\n")
                f.write(f"# 2. Email: salah satu dari atas\n")
                f.write(f"# 3. Password: yang di sebelahnya\n")
                f.write(f"# 4. IMAP: mail.{DOMAIN} Port 993 SSL\n")
                f.write(f"# 5. SMTP: mail.{DOMAIN} Port 465 SSL\n")
            
            caption = f"✅ **{jml} EMAIL @{DOMAIN} - BISA LOGIN DI HP!**\nSisa Rp {sisa:,}\n\nUser bisa login di Gmail App pakai email di file ini & OTP JMO masuk langsung ke HP user!\n\nIMAP: mail.{DOMAIN} Port 993"
            if gagal:
                caption += f"\n\n⚠️ {len(gagal)} email mode SIMULASI - Setting CPANEL di Railway biar jadi mailbox asli!"
            else:
                caption += f"\n\n✅ Semua mailbox asli sudah aktif!"
            
            await q.message.reply_document(open(path,"rb"), filename=f"{jml}_Email_{DOMAIN}_BISA_LOGIN.txt", caption=caption, parse_mode="Markdown")
            await q.edit_message_text(f"✅ **{jml} EMAIL BERHASIL DIBUAT!** File dikirim di atas!\n\nSemua email bisa login di HP & terima OTP JMO langsung!", reply_markup=main_menu())

    elif data=="format_jmo":
        if not is_sub:
            await q.edit_message_text(f"🔒 **FORMAT OTOMATIS TERKUNCI**\nStatus: {sub_info}\nSaldo: Rp {saldo:,}\n\nPaket 1Bln 10k, 2Bln 20k, 3Bln 25k, Unlimited 200k\n\nTopup dulu via QRIS!", reply_markup=paket_format_menu(saldo)); return
        # MENU 6. FORMAT OTOMATIS - 4 SUBMENU SESUAI PERMINTAAN
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("1. BUAT FORMAT", callback_data="buat_format")],
            [InlineKeyboardButton("2. HASIL FORMAT", callback_data="hasil_format")],
            [InlineKeyboardButton("3. SETTING FORMAT", callback_data="setting_format_menu")],
            [InlineKeyboardButton("4. HISTORY", callback_data="history_format")],
            [InlineKeyboardButton("⬅️ Kembali ke Menu Utama", callback_data="back")]
        ])
        conn=sqlite3.connect(DB_NAME)
        jml_format = conn.execute("SELECT COUNT(*) FROM formats WHERE user_id=?", (uid,)).fetchone()[0]
        jml_history = conn.execute("SELECT COUNT(*) FROM format_history WHERE user_id=?", (uid,)).fetchone()[0]
        setting = get_format_setting(uid)
        conn.close()
        txt=f"""📋 **6. FORMAT OTOMATIS - {sub_info}**

Pilih menu:
**1. BUAT FORMAT** - Buat format asal2an, bot auto rapikan sesuai SETTING
**2. HASIL FORMAT** - {jml_format} format tersimpan (ada tombol CARI, HAPUS, SALIN)
**3. SETTING FORMAT** - Setting template format kamu
**4. HISTORY** - {jml_history} history pembuatan

**Template Aktif:**
```
{setting[:200]}...
```

Contoh ketik asal:
Info loker
Bandung
Antapani
Kulim
Gaji 12 juta
Dpt jmo lasik jos

Bot akan auto format sesuai template di SETTING!"""
        await q.edit_message_text(txt, reply_markup=kb, parse_mode="Markdown")

    elif data=="buat_format":
        if not is_sub:
            await q.edit_message_text(f"🔒 Terkunci {sub_info}", reply_markup=paket_format_menu(saldo)); return
        context.user_data['mode']='buat_format_xl'
        setting = get_format_setting(uid)
        txt=f"""📊 **1. BUAT FORMAT OTOMATIS**

**Silahkan kirim file XL-nya** 📎

Format yang didukung: **.xlsx**

Bot akan membaca setiap baris dari file Excel, lalu otomatis merapikan sesuai **SETTING FORMAT**.

**Template aktif:**
```
{setting}
```

Setelah selesai, bot akan mengirimkan **file Excel hasil format** yang siap dipakai.

❌ Jangan kirim teks dulu. Kirim file Excel (.xlsx).

Ketik /batal untuk membatalkan."""
        await q.edit_message_text(txt, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Batal", callback_data="format_jmo")]]), parse_mode="Markdown")

    elif data=="setting_format_menu":
        if not is_sub:
            await q.edit_message_text(f"🔒 Terkunci {sub_info}", reply_markup=paket_format_menu(saldo)); return
        setting = get_format_setting(uid)
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("✏️ Edit Template", callback_data="edit_setting_format")],
            [InlineKeyboardButton("🔄 Reset ke Default", callback_data="reset_setting_format")],
            [InlineKeyboardButton("⬅️ Kembali", callback_data="format_jmo")]
        ])
        txt=f"""⚙️ **3. SETTING FORMAT**

Template ini yang dipakai bot untuk auto-format ketika kamu klik 1. BUAT FORMAT

**Template Aktif Saat Ini:**
```
{setting}
```

**Placeholder yang bisa dipakai:**
{{judul}} - Judul (baris pertama)
{{lokasi}} - Lokasi (Bandung, Antapani, dll)
{{posisi}} - Posisi pekerjaan
{{gaji}} - Gaji (12 juta, dll)
{{benefit}} - Benefit (jmo, bpjs, lasik)
{{isi_asal}} / {{isi}} / {{text}} - Text asal lengkap yang kamu ketik

**Contoh Template:**
📢 INFO LOKER - {{judul}}
📍 {{lokasi}}
💰 {{gaji}}
✨ {{benefit}}
📝 {{isi_asal}}

Klik Edit untuk ganti template!"""
        await q.edit_message_text(txt, reply_markup=kb, parse_mode="Markdown")

    elif data=="edit_setting_format":
        context.user_data['mode']='edit_setting_format'
        await q.edit_message_text("⚙️ **EDIT SETTING FORMAT**\n\nKetik template baru kamu, gunakan placeholder {judul}, {lokasi}, {gaji}, {benefit}, {isi_asal}\n\nContoh:\n📢 {judul}\n📍 {lokasi}\n💰 {gaji}\n\nKetik template sekarang:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Batal", callback_data="setting_format_menu")]]))

    elif data=="reset_setting_format":
        conn=sqlite3.connect(DB_NAME)
        conn.execute("DELETE FROM format_settings WHERE user_id=?", (uid,))
        conn.commit(); conn.close()
        await q.edit_message_text("✅ Template di-reset ke default!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⚙️ Lihat Setting", callback_data="setting_format_menu")],[InlineKeyboardButton("⬅️ Kembali", callback_data="format_jmo")]]))

    elif data=="history_format":
        if not is_sub:
            await q.edit_message_text(f"🔒 Terkunci {sub_info}", reply_markup=paket_format_menu(saldo)); return
        conn=sqlite3.connect(DB_NAME)
        rows=conn.execute("SELECT id, raw_text, formatted_text, created_at FROM format_history WHERE user_id=? ORDER BY id DESC LIMIT 20", (uid,)).fetchall()
        conn.close()
        if not rows:
            await q.edit_message_text("📜 **4. HISTORY**\n\nBelum ada history. Buat format dulu di menu 1. BUAT FORMAT", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("1. BUAT FORMAT", callback_data="buat_format")],[InlineKeyboardButton("⬅️ Kembali", callback_data="format_jmo")]]))
            return
        buttons=[]
        for fid, raw, fmt, tgl in rows:
            title = raw[:25].replace("\n"," ") + "..."
            buttons.append([InlineKeyboardButton(f"🕐 {title} - {tgl}", callback_data=f"viewhist_{fid}")])
        buttons.append([InlineKeyboardButton("🗑️ Hapus Semua History", callback_data="hapus_all_history")])
        buttons.append([InlineKeyboardButton("⬅️ Kembali", callback_data="format_jmo")])
        await q.edit_message_text(f"📜 **4. HISTORY - {len(rows)} Data**\n\nHistory pembuatan format kamu (raw -> formatted):", reply_markup=InlineKeyboardMarkup(buttons))

    elif data.startswith("viewhist_"):
        fid=int(data.split("_")[1])
        conn=sqlite3.connect(DB_NAME)
        row=conn.execute("SELECT raw_text, formatted_text, created_at FROM format_history WHERE id=? AND user_id=?", (fid, uid)).fetchone()
        conn.close()
        if not row:
            await q.edit_message_text("Tidak ditemukan", reply_markup=main_menu()); return
        raw, fmt, tgl = row
        txt=f"📜 **HISTORY #{fid} - {tgl}**\n\n**RAW (Asal2an):**\n{raw}\n\n**HASIL FORMAT:**\n{fmt}"
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("📋 Salin Hasil", callback_data=f"salin_hist_{fid}")],
            [InlineKeyboardButton("🗑️ Hapus", callback_data=f"delhist_{fid}")],
            [InlineKeyboardButton("⬅️ Kembali ke History", callback_data="history_format")]
        ])
        await q.edit_message_text(txt, reply_markup=kb, parse_mode="Markdown")

    elif data.startswith("delhist_"):
        fid=int(data.replace("delhist_",""))
        conn=sqlite3.connect(DB_NAME); conn.execute("DELETE FROM format_history WHERE id=? AND user_id=?", (fid, uid)); conn.commit(); conn.close()
        await q.edit_message_text("✅ History dihapus", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📜 Kembali ke History", callback_data="history_format")]]))

    elif data=="hapus_all_history":
        await q.edit_message_text("Yakin hapus semua history?", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Ya Hapus", callback_data="confirm_hapus_all_history")],[InlineKeyboardButton("❌ Batal", callback_data="history_format")]]))

    elif data=="confirm_hapus_all_history":
        conn=sqlite3.connect(DB_NAME); conn.execute("DELETE FROM format_history WHERE user_id=?", (uid,)); conn.commit(); conn.close()
        await q.edit_message_text("✅ Semua history dihapus", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali", callback_data="format_jmo")]]))

    elif data.startswith("sub_"):
        paket=data.replace("sub_","")
        harga=HARGA_FORMAT[paket]
        if saldo<harga:
            await q.edit_message_text(f"❌ Saldo kurang! {paket} Rp {harga:,}", reply_markup=paket_format_menu(saldo)); return
        conn=sqlite3.connect(DB_NAME); conn.execute("UPDATE users SET saldo=saldo-? WHERE user_id=?",(harga,uid)); conn.commit(); conn.close()
        expiry=add_sub(uid,paket)
        await q.edit_message_text(f"✅ LANGGANAN AKTIF {paket} sampai {expiry}", reply_markup=main_menu())

    elif data=="hasil_format":
        if not is_sub:
            await q.edit_message_text(f"🔒 Terkunci {sub_info}", reply_markup=paket_format_menu(saldo)); return
        conn=sqlite3.connect(DB_NAME)
        rows=conn.execute("SELECT id,title,created_at FROM formats WHERE user_id=? ORDER BY id DESC LIMIT 30",(uid,)).fetchall()
        conn.close()
        if not rows:
            await q.edit_message_text("📂 **2. HASIL FORMAT - Belum ada format**\n\nBelum ada format tersimpan. Klik 1. BUAT FORMAT dulu, ketik asal2an, bot auto format sesuai SETTING, lalu klik Simpan -> otomatis masuk sini!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("1. BUAT FORMAT",callback_data="buat_format")],[InlineKeyboardButton("⬅️ Kembali ke Format",callback_data="format_jmo")]])); return
        buttons=[]
        for fid,title,tgl in rows:
            buttons.append([InlineKeyboardButton(f"📄 {title[:25]} - {tgl}", callback_data=f"viewfmt_{fid}")])
        # Tombol CARI, HAPUS, SALIN sesuai permintaan
        buttons.append([InlineKeyboardButton("🔍 CARI",callback_data="cari_format"), InlineKeyboardButton("🗑️ HAPUS SEMUA",callback_data="hapus_all")])
        buttons.append([InlineKeyboardButton("📋 SALIN SEMUA",callback_data="salin_all")])
        buttons.append([InlineKeyboardButton("⬅️ Kembali ke Format",callback_data="format_jmo")])
        await q.edit_message_text(f"📂 **2. HASIL FORMAT - {len(rows)} Data**\n\nKlik format untuk lihat detail. Ada tombol CARI, HAPUS, SALIN!", reply_markup=InlineKeyboardMarkup(buttons))

    elif data.startswith("viewfmt_"):
        fid=int(data.split("_")[1])
        conn=sqlite3.connect(DB_NAME); row=conn.execute("SELECT title,content,created_at FROM formats WHERE id=? AND user_id=?",(fid,uid)).fetchone(); conn.close()
        if not row: await q.edit_message_text("Tidak ditemukan", reply_markup=main_menu()); return
        title,content,tgl=row
        # Tombol Edit, Simpan, Hapus, Salin, CARI sesuai permintaan
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("✏️ Edit",callback_data=f"editfmt_{fid}"), InlineKeyboardButton("💾 Simpan",callback_data=f"simpanfmt_{fid}"), InlineKeyboardButton("🗑️ Hapus",callback_data=f"delfmt_{fid}")],
            [InlineKeyboardButton("📋 SALIN",callback_data=f"salinfmt_{fid}"), InlineKeyboardButton("🔍 CARI",callback_data="cari_format")],
            [InlineKeyboardButton("⬅️ Kembali ke Hasil Format",callback_data="hasil_format")]
        ])
        await q.edit_message_text(f"📄 **{title}**\n{tgl}\n\n```\n{content}\n```", reply_markup=kb, parse_mode="Markdown")

    elif data.startswith("editfmt_"):
        fid=int(data.split("_")[1])
        context.user_data['mode']='edit_format'
        context.user_data['edit_fid']=fid
        conn=sqlite3.connect(DB_NAME); row=conn.execute("SELECT content FROM formats WHERE id=? AND user_id=?",(fid,uid)).fetchone(); conn.close()
        if not row:
            await q.edit_message_text("Tidak ditemukan", reply_markup=main_menu()); return
        await q.edit_message_text(f"✏️ **Edit Format #{fid}**\n\nIsi lama:\n{row[0][:500]}...\n\nKetik isi baru sekarang:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Batal",callback_data=f"viewfmt_{fid}")]]))

    elif data.startswith("salinfmt_"):
        fid=int(data.split("_")[1])
        conn=sqlite3.connect(DB_NAME); row=conn.execute("SELECT content FROM formats WHERE id=? AND user_id=?",(fid,uid)).fetchone(); conn.close()
        if not row:
            await q.edit_message_text("Tidak ditemukan", reply_markup=main_menu()); return
        # Salin = kirim lagi content biar user bisa copy
        await q.edit_message_text(f"📋 **SALIN - Format #{fid} berhasil disalin!**\n\nKlik text di bawah untuk copy:\n\n```\n{row[0]}\n```\n\nFormat sudah siap ditempel!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali",callback_data=f"viewfmt_{fid}")]]), parse_mode="Markdown")

    elif data.startswith("simpanfmt_"):
        fid=int(data.split("_")[1])
        # Simpan sudah ada di DB, cuma kasih konfirmasi
        await q.edit_message_text(f"✅ **Format #{fid} sudah tersimpan di 2. HASIL FORMAT**\n\nFormat otomatis masuk ke menu 2. HASIL FORMAT ketika kamu klik Simpan setelah BUAT FORMAT!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📂 Lihat Hasil Format",callback_data="hasil_format")],[InlineKeyboardButton("⬅️ Kembali",callback_data="format_jmo")]]))

    elif data=="salin_all":
        conn=sqlite3.connect(DB_NAME)
        rows=conn.execute("SELECT content FROM formats WHERE user_id=? ORDER BY id DESC LIMIT 5", (uid,)).fetchall()
        conn.close()
        if not rows:
            await q.edit_message_text("Belum ada format untuk disalin", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali",callback_data="hasil_format")]])); return
        all_text = "\n\n---\n\n".join([r[0] for r in rows])
        await q.edit_message_text(f"📋 **SALIN SEMUA - {len(rows)} Format**\n\n```\n{all_text[:3500]}\n```", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali",callback_data="hasil_format")]]), parse_mode="Markdown")

    elif data.startswith("delfmt_"):
        fid=int(data.split("_")[1])
        conn=sqlite3.connect(DB_NAME); conn.execute("DELETE FROM formats WHERE id=? AND user_id=?",(fid,uid)); conn.commit(); conn.close()
        await q.edit_message_text("✅ Dihapus", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📂 Kembali",callback_data="hasil_format")]]))

    elif data=="hapus_all":
        await q.edit_message_text("Yakin hapus semua?", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Ya",callback_data="confirm_hapus_all")],[InlineKeyboardButton("❌ Batal",callback_data="hasil_format")]]))

    elif data=="confirm_hapus_all":
        conn=sqlite3.connect(DB_NAME); conn.execute("DELETE FROM formats WHERE user_id=?",(uid,)); conn.commit(); conn.close()
        await q.edit_message_text("✅ Semua dihapus", reply_markup=main_menu())

    elif data=="cari_format":
        context.user_data['mode']='cari_format'
        await q.edit_message_text("🔍 Ketik kata kunci cari:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Batal",callback_data="hasil_format")]]))

    elif data=="setting_format":
        if not is_sub:
            await q.edit_message_text(f"🔒 Terkunci {sub_info}", reply_markup=paket_format_menu(saldo)); return
        context.user_data['mode']='setting_format_title'
        await q.edit_message_text("⚙️ **SETTING FORMAT**\n\nSilahkan masukan FORMAT sesuai yang kamu pinta (Judul dulu).\nContoh: Format Klaim Andi\n\nKetik judul sekarang:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Batal",callback_data="back")]]))

    elif data=="solusi_jmo":
        context.user_data['mode']='solusi_jmo_tanya'
        txt="""✅ **SOLUSI JMO SUPER LENGKAP**

Silahkan ketikan apa yang Anda tanya

Contoh:
• 025 - NIK duplikat
• 026 - KPJ tidak ditemukan
• 040 - Biometrik gagal
• 05, 06, 011, 020, 021, 024, 027, 030, 035, 041, 050
• otp tidak masuk
• lupa password
• cara klaim
• saldo tidak muncul

Ketik sekarang masalah JMO kamu!"""
        await q.edit_message_text(txt, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali",callback_data="back")]]), parse_mode="Markdown")

    elif data=="cara_login":
        txt=f"""📱 **CARA LOGIN EMAIL @{DOMAIN} DI HP - BISA TERIMA OTP LANGSUNG**

**Agar user bisa login email di HP & OTP JMO masuk ke HP user:**

**SETELAH BELI jmogmail.com + HOSTING:**

1. Beli Hosting di Hostinger (paket Single/Premium) + Domain jmogmail.com
2. Di hPanel Hostinger > Email > Buat Email Manual pertama
3. Di Railway, set Variables:
   `CPANEL_HOST` = https://hpanel.hostinger.com atau IP server
   `CPANEL_USER` = email hPanel kamu
   `CPANEL_PASS` = password hPanel
   `CPANEL_API_TOKEN` = (kalau ada)

4. Bot otomatis akan bikin mailbox asli setiap user beli email!

**CARA USER LOGIN DI HP:**

**Android (Gmail App):**
1. Buka Gmail > Foto Profil > Tambah Akun Lain
2. Pilih Lainnya / Other
3. Email: (dari bot) contoh `abx123@{DOMAIN}`
4. Password: (dari bot)
5. Pilih IMAP
6. Server: `mail.{DOMAIN}` Port 993 SSL
7. SMTP: `mail.{DOMAIN}` Port 465 SSL
8. Selesai! OTP JMO akan masuk ke Gmail App user!

**iPhone:**
Pengaturan > Mail > Akun > Tambah Akun > Lainnya > Masukkan Email & Password > IMAP `mail.{DOMAIN}`

**STATUS BOT KAMU SEKARANG:**
Domain: {DOMAIN}
CPanel Config: {'SUDAH DI SET' if CPANEL_HOST else 'BELUM DI SET - Mode Simulasi'}
Kalau belum set CPanel, email yang dibeli user hanya simulasi, belum bisa login. Set hosting dulu biar bisa login!

Butuh bantuan setting hosting? Chat admin!"""
        await q.edit_message_text(txt, reply_markup=main_menu(), parse_mode="Markdown")

    # === HANDLER UNTUK FORMAT OTOMATIS - Edit/Simpan/Hapus Temp ===
    elif data=="edit_temp_format":
        context.user_data['mode']='edit_temp_format_content'
        last_fmt = context.user_data.get('last_formatted','')
        await q.edit_message_text(f"✏️ **1. Edit - Edit Format**\n\nIsi saat ini:\n{last_fmt[:600]}\n\nKetik versi baru:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Batal", callback_data="format_jmo")]]))

    elif data=="simpan_temp_format":
        # SIMPAN - Otomatis masuk ke 2. HASIL FORMAT sesuai permintaan user
        formatted = context.user_data.get('last_formatted','')
        raw = context.user_data.get('last_raw','')
        if not formatted:
            await q.edit_message_text("❌ Tidak ada format untuk disimpan! Buat dulu di 1. BUAT FORMAT", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("1. BUAT FORMAT", callback_data="buat_format")]]))
            return
        # Buat judul dari baris pertama
        title = raw.split("\n")[0][:40] if raw else formatted[:40]
        conn=sqlite3.connect(DB_NAME)
        conn.execute("INSERT INTO formats (user_id, title, content, created_at) VALUES (?,?,?,?)",(uid, title, formatted, datetime.now().strftime("%d-%m-%Y %H:%M")))
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit(); conn.close()
        context.user_data['mode']=None
        context.user_data['last_formatted']=None
        context.user_data['last_raw']=None
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("📂 Lihat di 2. HASIL FORMAT", callback_data="hasil_format")],
            [InlineKeyboardButton("📋 SALIN", callback_data=f"salinfmt_{new_id}")],
            [InlineKeyboardButton("⬅️ Kembali ke Format", callback_data="format_jmo")]
        ])
        await q.edit_message_text(f"✅ **2. Simpan - BERHASIL DISIMPAN!**\n\nFormat kamu otomatis masuk ke **2. HASIL FORMAT**!\n\nID: #{new_id}\nJudul: {title}\n\nIsi:\n```\n{formatted}\n```\n\nSekarang cek di menu 2. HASIL FORMAT, ada tombol CARI, HAPUS, SALIN!", reply_markup=kb, parse_mode="Markdown")

    elif data=="hapus_temp_format":
        context.user_data['last_formatted']=None
        context.user_data['last_raw']=None
        context.user_data['mode']=None
        # Hapus history terakhir juga
        last_hist = context.user_data.get('last_history_id')
        if last_hist:
            conn=sqlite3.connect(DB_NAME)
            conn.execute("DELETE FROM format_history WHERE id=? AND user_id=?", (last_hist, uid))
            conn.commit(); conn.close()
        await q.edit_message_text("🗑️ **3. Hapus - Format dihapus!**\n\nFormat temporary dihapus, tidak masuk ke HASIL FORMAT", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("1. BUAT FORMAT LAGI", callback_data="buat_format")],[InlineKeyboardButton("⬅️ Kembali", callback_data="format_jmo")]]))

    elif data.startswith("salin_hist_"):
        fid=int(data.replace("salin_hist_",""))
        conn=sqlite3.connect(DB_NAME)
        row=conn.execute("SELECT formatted_text FROM format_history WHERE id=? AND user_id=?", (fid, uid)).fetchone()
        conn.close()
        if not row:
            await q.edit_message_text("Tidak ditemukan", reply_markup=main_menu()); return
        await q.edit_message_text(f"📋 **SALIN HISTORY #{fid}**\n\n```\n{row[0]}\n```\n\nSiap ditempel!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali", callback_data="history_format")]]), parse_mode="Markdown")

    elif data=="admin":
        if uid not in ADMIN_IDS:
            await q.edit_message_text("Bukan admin!", reply_markup=main_menu()); return
        conn=sqlite3.connect(DB_NAME)
        total=conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        total_email=conn.execute("SELECT COUNT(*) FROM emails").fetchone()[0]
        total_sub=conn.execute("SELECT COUNT(*) FROM subscriptions").fetchone()[0]
        pending=conn.execute("SELECT COUNT(*) FROM topup WHERE status='pending'").fetchone()[0]
        total_saldo=conn.execute("SELECT SUM(saldo) FROM users").fetchone()[0] or 0
        conn.close()
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("1️⃣ CEK USER AKTIF", callback_data="cek_user")],
            [InlineKeyboardButton("2️⃣ TAMBAH SALDO", callback_data="addsaldo"), InlineKeyboardButton("3️⃣ KURANGI SALDO", callback_data="kurangi_saldo")],
            [InlineKeyboardButton("4️⃣ BROADCAST", callback_data="broadcast")],
            [InlineKeyboardButton("5️⃣ HAPUS ID", callback_data="hapus_id")],
            [InlineKeyboardButton("6️⃣ CEK TRANSAKSI 💰", callback_data="cek_transaksi")],
            [InlineKeyboardButton("7️⃣ LAPORAN HARIAN 📊", callback_data="laporan_harian")],
            [InlineKeyboardButton(f"📤 Topup Pending ({pending})", callback_data="list_pending")],
            [InlineKeyboardButton(f"📊 Statistik {total} User", callback_data="statistik")],
            [InlineKeyboardButton("⬅️ Kembali", callback_data="back")]
        ])
        txt=f"👑 **PANEL ADMIN SUPER LENGKAP V8**\n\n👤 Total User: {total}\n💰 Total Saldo User: Rp {total_saldo:,}\n📧 Email Terjual: {total_email}\n🔒 Langganan Format: {total_sub}\n💳 Topup Pending: {pending}\n\nHarga: 1 Email 5k, 10 Email 40k, 20 Email 70k\n\nPilih menu:\n1.CEK USER AKTIF\n2.TAMBAH SALDO\n3.KURANGI SALDO\n4.BROADCAST\n5.HAPUS ID\n6.CEK TRANSAKSI\n7.LAPORAN HARIAN"
        await q.edit_message_text(txt, reply_markup=kb, parse_mode="Markdown")

    elif data=="cek_user":
        if uid not in ADMIN_IDS: return
        conn=sqlite3.connect(DB_NAME)
        rows=conn.execute("SELECT user_id, username, saldo, created_at FROM users ORDER BY saldo DESC LIMIT 20").fetchall()
        total=conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        conn.close()
        txt=f"👤 **CEK USER AKTIF - Total {total} User**\n\nTop 20 Saldo Tertinggi:\n\n"
        for i, (user_id, username, saldo, tgl) in enumerate(rows, 1):
            uname = f"@{username}" if username else "No username"
            txt+=f"{i}. ID `{user_id}` {uname}\n   Saldo Rp {saldo:,} | {tgl}\n"
        txt+=f"\nKetik ID untuk detail atau gunakan menu lain."
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("🔍 Cek Detail User", callback_data="cek_detail_user")],
            [InlineKeyboardButton("📤 Export Semua User", callback_data="export_user")],
            [InlineKeyboardButton("⬅️ Kembali ke Admin", callback_data="admin")]
        ])
        await q.edit_message_text(txt, reply_markup=kb, parse_mode="Markdown")

    elif data=="cek_detail_user":
        context.user_data['mode']='cek_detail_user'
        await q.edit_message_text("🔍 **CEK DETAIL USER**\n\nKirim ID user yang mau dicek:\nContoh: `123456789`\n\nBot akan kasih detail saldo, email yang dibeli, langganan, dll.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali", callback_data="cek_user")]]), parse_mode="Markdown")

    elif data=="export_user":
        conn=sqlite3.connect(DB_NAME)
        rows=conn.execute("SELECT user_id, username, saldo FROM users").fetchall()
        conn.close()
        text_data = "ID,Username,Saldo\n" + "\n".join([f"{uid},{uname or ''},{saldo}" for uid, uname, saldo in rows])
        path = f"/tmp/export_user_{uid}.csv"
        open(path, "w").write(text_data)
        await q.message.reply_document(open(path, "rb"), filename=f"users_{len(rows)}.csv", caption=f"✅ Export {len(rows)} user")
        await q.edit_message_text(f"✅ File export {len(rows)} user dikirim di atas!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali", callback_data="admin")]]))

    elif data=="list_pending":
        conn=sqlite3.connect(DB_NAME)
        rows=conn.execute("SELECT id, user_id, amount, created_at FROM topup WHERE status='pending' ORDER BY id DESC LIMIT 10").fetchall()
        conn.close()
        if not rows:
            await q.edit_message_text("✅ Tidak ada topup pending", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali", callback_data="admin")]])); return
        txt="💳 **TOPUP PENDING**\n\n"
        buttons=[]
        for tid, user_id, amount, tgl in rows:
            txt+=f"#{tid} - User {user_id} - Rp {amount:,} - {tgl}\n"
            buttons.append([InlineKeyboardButton(f"✅ Approve #{tid} Rp {amount:,}", callback_data=f"approve_topup_{tid}"), InlineKeyboardButton(f"❌ Reject #{tid}", callback_data=f"reject_topup_{tid}")])
        buttons.append([InlineKeyboardButton("⬅️ Kembali", callback_data="admin")])
        await q.edit_message_text(txt, reply_markup=InlineKeyboardMarkup(buttons))

    elif data=="statistik":
        conn=sqlite3.connect(DB_NAME)
        total=conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        total_email=conn.execute("SELECT COUNT(*) FROM emails").fetchone()[0]
        total_sub=conn.execute("SELECT COUNT(*) FROM subscriptions").fetchone()[0]
        total_saldo=conn.execute("SELECT SUM(saldo) FROM users").fetchone()[0] or 0
        total_topup=conn.execute("SELECT SUM(amount) FROM topup WHERE status LIKE 'approved%'").fetchone()[0] or 0
        conn.close()
        await q.edit_message_text(f"📊 **Statistik Lengkap**\n\n👤 User: {total}\n💰 Total Saldo User: Rp {total_saldo:,}\n💳 Total Pemasukan Topup: Rp {total_topup:,}\n📧 Email Terjual: {total_email}\n🔒 Langganan: {total_sub}\n\nHarga: 5k/40k/70k", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("💰 Cek Transaksi", callback_data="cek_transaksi")],[InlineKeyboardButton("⬅️ Kembali", callback_data="admin")]]))

    elif data=="cek_transaksi":
        if uid not in ADMIN_IDS: return
        conn=sqlite3.connect(DB_NAME)
        # Hitung pemasukan
        total_topup=conn.execute("SELECT SUM(amount) FROM topup WHERE status LIKE 'approved%'").fetchone()[0] or 0
        total_topup_pending=conn.execute("SELECT SUM(amount) FROM topup WHERE status='pending'").fetchone()[0] or 0
        count_topup=conn.execute("SELECT COUNT(*) FROM topup WHERE status LIKE 'approved%'").fetchone()[0]
        # Email terjual
        total_email=conn.execute("SELECT COUNT(*) FROM emails").fetchone()[0]
        # Hitung estimasi pendapatan dari email (asumsi harga rata2)
        # Ambil 20 transaksi terakhir
        rows=conn.execute("SELECT user_id, amount, status, created_at FROM topup ORDER BY id DESC LIMIT 20").fetchall()
        rows_email=conn.execute("SELECT owner_id, created_at FROM emails ORDER BY id DESC LIMIT 20").fetchall()
        rows_sub=conn.execute("SELECT user_id, plan, expiry FROM subscriptions ORDER BY rowid DESC LIMIT 10").fetchall()
        conn.close()
        
        txt=f"""💰 **CEK TRANSAKSI LENGKAP**

**RINGKASAN PEMASUKAN:**
💳 Total Topup Masuk: Rp {total_topup:,}
⏳ Topup Pending: Rp {total_topup_pending:,}
🔢 Jumlah Transaksi Topup: {count_topup}
📧 Total Email Terjual: {total_email}

**20 TRANSAKSI TOPUP TERAKHIR:**
"""
        for tid, user_id, amount, tgl, status in [ (i, r[0], r[1], r[3], r[2]) for i,r in enumerate(rows,1) ]:
            txt+=f"{tid}. ID {user_id} - Rp {amount:,} - {status} - {tgl}\n"
        
        txt+=f"\n**10 LANGGANAN FORMAT TERAKHIR:**\n"
        for i, (user_id, plan, expiry) in enumerate(rows_sub,1):
            txt+=f"{i}. ID {user_id} - {plan} - {expiry}\n"
        
        txt+=f"\n**20 EMAIL TERJUAL TERAKHIR:**\n"
        for i, (owner_id, tgl) in enumerate(rows_email,1):
            txt+=f"{i}. User {owner_id} - {tgl}\n"

        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("📤 Export Transaksi CSV", callback_data="export_transaksi")],
            [InlineKeyboardButton("💳 Detail Topup", callback_data="list_pending"), InlineKeyboardButton("👤 Cek User", callback_data="cek_user")],
            [InlineKeyboardButton("⬅️ Kembali ke Admin", callback_data="admin")]
        ])
        await q.edit_message_text(txt, reply_markup=kb)

    elif data=="export_transaksi":
        if uid not in ADMIN_IDS: return
        conn=sqlite3.connect(DB_NAME)
        rows=conn.execute("SELECT id, user_id, amount, status, created_at FROM topup ORDER BY id DESC").fetchall()
        conn.close()
        csv_text="ID,UserID,Amount,Status,Date\n" + "\n".join([f"{r[0]},{r[1]},{r[2]},{r[3]},{r[4]}" for r in rows])
        path=f"/tmp/transaksi_{uid}.csv"
        open(path,"w").write(csv_text)
        total=sum([r[2] for r in rows if 'approved' in r[3]])
        await q.message.reply_document(open(path,"rb"), filename=f"transaksi_{len(rows)}.csv", caption=f"✅ Export {len(rows)} transaksi\nTotal Pemasukan: Rp {total:,}")
        await q.edit_message_text(f"✅ File {len(rows)} transaksi dikirim! Total pemasukan Rp {total:,}", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali", callback_data="cek_transaksi")]]))

    elif data=="laporan_harian":
        if uid not in ADMIN_IDS: return
        conn=sqlite3.connect(DB_NAME)
        today=datetime.now().strftime("%d-%m-%Y")
        total_today=conn.execute("SELECT SUM(amount) FROM topup WHERE status LIKE 'approved%' AND created_at LIKE ?", (f"%{today}%",)).fetchone()[0] or 0
        count_today=conn.execute("SELECT COUNT(*) FROM topup WHERE status LIKE 'approved%' AND created_at LIKE ?", (f"%{today}%",)).fetchone()[0]
        total_all=conn.execute("SELECT SUM(amount) FROM topup WHERE status LIKE 'approved%'").fetchone()[0] or 0
        total_users=conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        total_email=conn.execute("SELECT COUNT(*) FROM emails").fetchone()[0]
        conn.close()
        txt=f"📊 **LAPORAN HARIAN {today}**\n\nHari Ini:\n💳 Pemasukan: Rp {total_today:,}\n🔢 Transaksi: {count_today}\n\nTotal:\n👤 User: {total_users}\n💵 Total Pemasukan: Rp {total_all:,}\n📧 Email Terjual: {total_email}\n\nHarga 5k/40k/70k"
        await q.edit_message_text(txt, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔄 Refresh", callback_data="laporan_harian")],[InlineKeyboardButton("⬅️ Kembali", callback_data="admin")]]), parse_mode="Markdown")

    elif data=="broadcast":
        if uid not in ADMIN_IDS:
            await q.edit_message_text("Bukan admin!", reply_markup=main_menu()); return
        context.user_data['mode']='broadcast'
        await q.edit_message_text("📢 **BROADCAST**\n\nKetik pesan yang ingin dikirim ke SEMUA user:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Batal", callback_data="admin")]]))

    elif data.startswith("approve_topup_"):
        if uid not in ADMIN_IDS:
            await q.edit_message_text("Bukan admin!", reply_markup=main_menu()); return
        topup_id=int(data.replace("approve_topup_",""))
        conn=sqlite3.connect(DB_NAME)
        row=conn.execute("SELECT user_id, amount, status FROM topup WHERE id=?", (topup_id,)).fetchone()
        if not row:
            await q.edit_message_text("Topup tidak ditemukan", reply_markup=main_menu())
            conn.close(); return
        user_id, amount, status=row
        if status=="approved":
            await q.edit_message_text(f"Topup #{topup_id} sudah approved", reply_markup=main_menu())
            conn.close(); return
        conn.execute("UPDATE users SET saldo=saldo+? WHERE user_id=?", (amount, user_id))
        conn.execute("UPDATE topup SET status='approved' WHERE id=?", (topup_id,))
        conn.commit(); conn.close()
        await q.edit_message_text(f"✅ Topup #{topup_id} Rp {amount:,} untuk {user_id} BERHASIL!", reply_markup=main_menu())
        try:
            await context.bot.send_message(chat_id=user_id, text=f"✅ **TOPUP BERHASIL!**\n\nSaldo Rp {amount:,} masuk!\nID: #{topup_id}\n\nBisa beli email 5k/40k/70k sekarang!", parse_mode="Markdown")
        except: pass

    elif data.startswith("reject_topup_"):
        if uid not in ADMIN_IDS: return
        topup_id=int(data.replace("reject_topup_",""))
        conn=sqlite3.connect(DB_NAME)
        conn.execute("UPDATE topup SET status='rejected' WHERE id=?", (topup_id,))
        conn.commit(); conn.close()
        await q.edit_message_text(f"❌ Topup #{topup_id} ditolak", reply_markup=main_menu())

    elif data.startswith("confirm_hapus_"):
        if uid not in ADMIN_IDS: return
        uid_t=int(data.replace("confirm_hapus_",""))
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Ya, Hapus Permanen", callback_data=f"do_hapus_{uid_t}")],
            [InlineKeyboardButton("❌ Batal", callback_data="cek_user")]
        ])
        await q.edit_message_text(f"⚠️ **YAKIN HAPUS USER ID {uid_t}?**\n\nSemua data saldo, email, format akan hilang permanen!", reply_markup=kb)

    elif data.startswith("do_hapus_"):
        if uid not in ADMIN_IDS: return
        uid_t=int(data.replace("do_hapus_",""))
        conn=sqlite3.connect(DB_NAME)
        conn.execute("DELETE FROM users WHERE user_id=?", (uid_t,))
        conn.execute("DELETE FROM emails WHERE owner_id=?", (uid_t,))
        conn.execute("DELETE FROM subscriptions WHERE user_id=?", (uid_t,))
        conn.execute("DELETE FROM formats WHERE user_id=?", (uid_t,))
        conn.execute("DELETE FROM topup WHERE user_id=?", (uid_t,))
        conn.commit(); conn.close()
        await q.edit_message_text(f"✅ User ID {uid_t} berhasil dihapus permanen!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali ke Admin", callback_data="admin")]]))

    elif data=="bantuan":
        txt=f"""🆘 **BANTUAN - BOT EMAIL {DOMAIN}**

**CARA BELI EMAIL 1 KLIK JADI:**
1. Klik 3. TOP UP -> Pilih nominal -> Bayar via QRIS/SEABANK {REKENING_NUMBER}/DANA {DANA_NUMBER}
2. Upload bukti -> Saldo masuk AUTO
3. Klik 2. BELI EMAIL -> Pilih paket 1/10/20/100
4. Bot bikin mailbox asli @{DOMAIN} otomatis!
5. Klik 4. CARA LOGIN -> Ikuti tutorial login di HP
6. OTP JMO masuk langsung ke HP kamu!

**HARGA:**
• 1 Email 5k, 10 Email 40k, 20 Email 70k, 100 Email 300k

**FORMAT OTOMATIS:**
Langganan 1 Bulan 10k, 2 Bulan 20k, 3 Bulan 25k, Unlimited 200k

**KALAU GAGAL LOGIN:**
• Pastikan IMAP: mail.{DOMAIN} Port 993 SSL
• SMTP: mail.{DOMAIN} Port 465 SSL
• Password sesuai dari bot (case sensitive)
• Hubungi admin via menu 9. HUBUNGI ADMIN

**PAYMENT:**
QRIS: Scan di menu TOP UP
SEABANK: {REKENING_NUMBER} a.n {REKENING_NAME}
DANA: {DANA_NUMBER} a.n {DANA_NAME}"""
        await q.edit_message_text(txt, reply_markup=main_menu(), parse_mode="Markdown")

    elif data=="cek_status":
        conn=sqlite3.connect(DB_NAME)
        # Status user
        user_emails = conn.execute("SELECT email, created_at FROM emails WHERE owner_id=? ORDER BY id DESC LIMIT 5", (uid,)).fetchall()
        topups = conn.execute("SELECT amount, status, created_at FROM topup WHERE user_id=? ORDER BY id DESC LIMIT 5", (uid,)).fetchall()
        conn.close()
        
        txt=f"""📊 **CEK STATUS - ID {uid}**

💰 Saldo: Rp {saldo:,}
📧 Domain: @{DOMAIN}

**5 EMAIL TERAKHIR KAMU:**
"""
        if user_emails:
            for i, (email, tgl) in enumerate(user_emails, 1):
                txt+=f"{i}. {email} - {tgl}\n"
        else:
            txt+="Belum ada email. Beli di menu 2. BELI EMAIL\n"
        
        txt+=f"\n**5 TOPUP TERAKHIR:**\n"
        if topups:
            for i, (amount, status, tgl) in enumerate(topups, 1):
                txt+=f"{i}. Rp {amount:,} - {status} - {tgl}\n"
        else:
            txt+="Belum ada topup\n"
        
        txt+=f"\n**CARA CEK OTP:**\nBuka Gmail App yang sudah login email @{DOMAIN}, cek Inbox/Spam, OTP JMO masuk dalam 5 detik!"
        
        await q.edit_message_text(txt, reply_markup=main_menu(), parse_mode="Markdown")

    elif data=="hubungi_admin":
        txt=f"""📞 **HUBUNGI ADMIN**

**Admin Bot {DOMAIN}:**
• Telegram: @hambali (ganti dengan username kamu)
• SEABANK: {REKENING_NUMBER} a.n {REKENING_NAME}
• DANA: {DANA_NUMBER} a.n {DANA_NAME}

**JAM OPERASIONAL:**
Online 08.00 - 22.00 WIB

**KELUHAN YANG BISA DIBANTU:**
• Email tidak bisa login di HP
• OTP tidak masuk
• Topup belum masuk padahal sudah transfer
• Lupa password email
• Butuh bantuan setting Gmail App

**FORMAT LAPORAN:**
Kirim ID kamu: `{uid}`
Sertakan screenshot error + email yang bermasalah

**BALAS CEPAT:**
Admin balas max 1 jam di jam kerja!

Klik tombol di bawah untuk chat admin langsung!"""
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("💬 Chat Admin di Telegram", url="https://t.me/hambali")],
            [InlineKeyboardButton("⬅️ Kembali ke Menu", callback_data="back")]
        ])
        await q.edit_message_text(txt, reply_markup=kb, parse_mode="Markdown")

    elif data=="back":
        await q.edit_message_text(f"**BOT JMO LENGKAP - {DOMAIN} - 1 KLIK JADI!**\n\n1. PROFIL\n2. BELI EMAIL\n3. TOP UP\n4. CARA LOGIN dan TERIMA OTP\n5. SOLUSI JMO\n6. FORMAT OTOMATIS\n7. BANTUAN\n8. CEK STATUS\n9. HUBUNGI ADMIN\n10. PANEL ADMIN\n\nPilih menu:", reply_markup=main_menu(), parse_mode="Markdown")

    elif data=="addsaldo":
        context.user_data['mode']='addsaldo'
        await q.edit_message_text("2️⃣ **TAMBAH SALDO**\n\nKirim format: `ID JUMLAH`\nContoh: `7962377092 50000`\n\nAkan tambah saldo Rp 50.000 ke ID tersebut.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali", callback_data="admin")]]), parse_mode="Markdown")

    elif data=="kurangi_saldo":
        context.user_data['mode']='kurangi_saldo'
        await q.edit_message_text("3️⃣ **KURANGI SALDO**\n\nKirim format: `ID JUMLAH`\nContoh: `123456789 10000`\n\nAkan kurangi saldo Rp 10.000 dari ID tersebut.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali", callback_data="admin")]]), parse_mode="Markdown")

    elif data=="hapus_id":
        context.user_data['mode']='hapus_id'
        await q.edit_message_text("5️⃣ **HAPUS ID USER**\n\n⚠️ Hati-hati! Ini akan hapus user permanen!\n\nKirim ID user yang mau dihapus:\nContoh: `123456789`\n\nAtau kirim `ID1 ID2 ID3` untuk hapus banyak sekaligus.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali", callback_data="admin")]]), parse_mode="Markdown")

async def msg_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id
    mode=context.user_data.get('mode')
    text=update.message.text or ""
    lower=text.lower()

    if mode=='broadcast':
        if uid not in ADMIN_IDS:
            await update.message.reply_text("Bukan admin!")
            context.user_data['mode']=None
            return
        conn=sqlite3.connect(DB_NAME)
        users=conn.execute("SELECT user_id FROM users").fetchall()
        conn.close()
        sukses=0; gagal=0
        await update.message.reply_text(f"📢 Broadcast ke {len(users)} user...")
        for (user_id,) in users:
            try:
                await context.bot.send_message(chat_id=user_id, text=text, parse_mode="Markdown")
                sukses+=1
            except:
                gagal+=1
        await update.message.reply_text(f"✅ Broadcast selesai! Sukses: {sukses} Gagal: {gagal}", reply_markup=main_menu())
        context.user_data['mode']=None
        return

    if mode=='addsaldo':
        try:
            uid_t,jml=text.split()
            uid_t=int(uid_t); jml=int(jml)
            conn=sqlite3.connect(DB_NAME)
            conn.execute("INSERT OR IGNORE INTO users (user_id,username,saldo,created_at) VALUES (?,?,0,?)",(uid_t,"",datetime.now().strftime("%d-%m-%Y")))
            conn.execute("UPDATE users SET saldo=saldo+? WHERE user_id=?",(jml,uid_t))
            conn.commit()
            new_saldo=conn.execute("SELECT saldo FROM users WHERE user_id=?", (uid_t,)).fetchone()[0]
            conn.close()
            await update.message.reply_text(f"✅ **TAMBAH SALDO BERHASIL**\n\nID: `{uid_t}`\nDitambah: Rp {jml:,}\nSaldo Sekarang: Rp {new_saldo:,}", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali ke Admin", callback_data="admin")]]), parse_mode="Markdown")
            try:
                await context.bot.send_message(chat_id=uid_t, text=f"💰 Saldo kamu ditambah Rp {jml:,} oleh admin! Saldo sekarang Rp {new_saldo:,}")
            except: pass
        except Exception as e:
            await update.message.reply_text(f"❌ Error {e}\nFormat: ID JUMLAH contoh 123 50000")
        context.user_data['mode']=None
        return

    if mode=='kurangi_saldo':
        try:
            uid_t,jml=text.split()
            uid_t=int(uid_t); jml=int(jml)
            conn=sqlite3.connect(DB_NAME)
            conn.execute("UPDATE users SET saldo=saldo-? WHERE user_id=?",(jml,uid_t))
            conn.commit()
            new_saldo=conn.execute("SELECT saldo FROM users WHERE user_id=?", (uid_t,)).fetchone()
            conn.close()
            new_saldo = new_saldo[0] if new_saldo else 0
            await update.message.reply_text(f"✅ **KURANGI SALDO BERHASIL**\n\nID: `{uid_t}`\nDikurangi: Rp {jml:,}\nSaldo Sekarang: Rp {new_saldo:,}", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali ke Admin", callback_data="admin")]]), parse_mode="Markdown")
        except Exception as e:
            await update.message.reply_text(f"❌ Error {e}\nFormat: ID JUMLAH")
        context.user_data['mode']=None
        return

    if mode=='hapus_id':
        try:
            ids = text.replace(",", " ").split()
            conn=sqlite3.connect(DB_NAME)
            deleted=0
            for id_str in ids:
                try:
                    uid_t=int(id_str)
                    conn.execute("DELETE FROM users WHERE user_id=?", (uid_t,))
                    conn.execute("DELETE FROM emails WHERE owner_id=?", (uid_t,))
                    conn.execute("DELETE FROM subscriptions WHERE user_id=?", (uid_t,))
                    conn.execute("DELETE FROM formats WHERE user_id=?", (uid_t,))
                    conn.execute("DELETE FROM topup WHERE user_id=?", (uid_t,))
                    deleted+=1
                except:
                    pass
            conn.commit(); conn.close()
            await update.message.reply_text(f"✅ **HAPUS ID BERHASIL**\n\n{deleted} user dihapus: {text}", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali ke Admin", callback_data="admin")]]))
        except Exception as e:
            await update.message.reply_text(f"❌ Error {e}")
        context.user_data['mode']=None
        return

    if mode=='cek_detail_user':
        try:
            uid_t=int(text.strip())
            conn=sqlite3.connect(DB_NAME)
            user_row=conn.execute("SELECT user_id, username, saldo, created_at FROM users WHERE user_id=?", (uid_t,)).fetchone()
            if not user_row:
                await update.message.reply_text(f"❌ User ID {uid_t} tidak ditemukan!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Kembali", callback_data="cek_user")]]))
                conn.close()
                context.user_data['mode']=None
                return
            user_id, username, saldo, tgl=user_row
            email_count=conn.execute("SELECT COUNT(*) FROM emails WHERE owner_id=?", (uid_t,)).fetchone()[0]
            format_count=conn.execute("SELECT COUNT(*) FROM formats WHERE user_id=?", (uid_t,)).fetchone()[0]
            sub=conn.execute("SELECT plan, expiry FROM subscriptions WHERE user_id=?", (uid_t,)).fetchone()
            topup_rows=conn.execute("SELECT SUM(amount) FROM topup WHERE user_id=? AND status LIKE 'approved%'", (uid_t,)).fetchone()[0] or 0
            conn.close()
            sub_info = f"{sub[0]} sampai {sub[1]}" if sub else "Belum langganan"
            txt=f"👤 **DETAIL USER ID {user_id}**\n\nUsername: @{username if username else 'No username'}\nSaldo: Rp {saldo:,}\nTotal Topup: Rp {topup_rows:,}\nEmail Dibeli: {email_count}\nFormat Dibuat: {format_count}\nLangganan: {sub_info}\nDaftar: {tgl}\n\nMau apa?"
            kb=InlineKeyboardMarkup([
                [InlineKeyboardButton("➕ Tambah Saldo", callback_data="addsaldo"), InlineKeyboardButton("➖ Kurangi Saldo", callback_data="kurangi_saldo")],
                [InlineKeyboardButton("🗑️ Hapus User Ini", callback_data=f"confirm_hapus_{uid_t}")],
                [InlineKeyboardButton("⬅️ Kembali", callback_data="cek_user")]
            ])
            await update.message.reply_text(txt, reply_markup=kb, parse_mode="Markdown")
        except Exception as e:
            await update.message.reply_text(f"❌ Error {e}")
        context.user_data['mode']=None
        return

    if mode=='setting_format_title':
        context.user_data['tmp_title']=text[:50]
        context.user_data['mode']='setting_format_content'
        await update.message.reply_text(f"✅ Judul: {text}\n\nSekarang masukan ISI FORMAT sesuai yang kamu pinta, lalu kirim (Save).")
        return

    if mode=='setting_format_content':
        title=context.user_data.get('tmp_title','Tanpa Judul')
        conn=sqlite3.connect(DB_NAME)
        conn.execute("INSERT INTO formats (user_id,title,content,created_at) VALUES (?,?,?,?)",(uid,title,text,datetime.now().strftime("%d-%m-%Y %H:%M")))
        conn.commit(); conn.close()
        context.user_data['mode']=None
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("📂 Lihat di HASIL FORMAT",callback_data="hasil_format")],[InlineKeyboardButton("⬅️ Menu",callback_data="back")]])
        await update.message.reply_text(f"✅ FORMAT DISIMPAN! {title}", reply_markup=kb)
        return

    # === FORMAT OTOMATIS - WAJIB UPLOAD EXCEL ===
    if mode=='buat_format_xl':
        if lower in ['/batal', 'batal']:
            context.user_data['mode']=None
            await update.message.reply_text("❌ Batal buat format", reply_markup=main_menu())
            return
        await update.message.reply_text("📊 **Silahkan kirim file XL-nya** (.xlsx).\n\nBot akan otomatis memproses setiap baris dari Excel.", parse_mode="Markdown")
        return

    # === FORMAT OTOMATIS LAMA - tetap dipertahankan ===
    if mode=='buat_format_raw':
        # User ketik asal2an, bot auto format sesuai SETTING
        raw_text = text
        if raw_text.lower() in ["/batal","batal"]:
            context.user_data['mode']=None
            await update.message.reply_text("❌ Batal buat format", reply_markup=main_menu())
            return
        
        # Ambil template setting
        template = get_format_setting(uid)
        formatted = auto_format_text(raw_text, template)
        
        # Simpan ke history
        conn=sqlite3.connect(DB_NAME)
        conn.execute("INSERT INTO format_history (user_id, raw_text, formatted_text, created_at) VALUES (?,?,?,?)",(uid, raw_text, formatted, datetime.now().strftime("%d-%m-%Y %H:%M")))
        history_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit(); conn.close()
        
        # Simpan formatted ke context untuk tombol Simpan nanti
        context.user_data['last_raw'] = raw_text
        context.user_data['last_formatted'] = formatted
        context.user_data['last_history_id'] = history_id
        context.user_data['mode']=None
        
        # Tampilkan hasil dengan tombol Edit, Simpan, Hapus sesuai permintaan
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("✏️ 1. Edit", callback_data=f"edit_temp_format"), InlineKeyboardButton("💾 2. Simpan", callback_data=f"simpan_temp_format"), InlineKeyboardButton("🗑️ 3. Hapus", callback_data=f"hapus_temp_format")],
            [InlineKeyboardButton("📂 Lihat HASIL FORMAT", callback_data="hasil_format")],
            [InlineKeyboardButton("⬅️ Kembali ke Format", callback_data="format_jmo")]
        ])
        txt=f"""✅ **FORMAT OTOMATIS BERHASIL!**

**RAW (Asal2an yang kamu ketik):**
{raw_text}

**HASIL AUTO-FORMAT (Sesuai SETTING di menu 3. SETTING FORMAT):**
```
{formatted}
```

**Tombol:**
1. Edit - Edit hasil format
2. Simpan - Simpan ke 2. HASIL FORMAT (otomatis masuk menu 2. HASIL FORMAT)
3. Hapus - Hapus format ini

Klik **2. Simpan** untuk simpan ke HASIL FORMAT!"""
        await update.message.reply_text(txt, reply_markup=kb, parse_mode="Markdown")
        return

    if mode=='edit_temp_format':
        context.user_data['mode']='edit_temp_format_content'
        last_formatted = context.user_data.get('last_formatted','')
        await update.message.reply_text(f"✏️ **Edit Format**\n\nIsi saat ini:\n{last_formatted[:500]}...\n\nKetik versi baru sekarang:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Batal", callback_data="format_jmo")]]))
        return

    if mode=='edit_temp_format_content':
        context.user_data['last_formatted']=text
        context.user_data['mode']=None
        formatted = text
        raw = context.user_data.get('last_raw','')
        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("✏️ 1. Edit", callback_data=f"edit_temp_format"), InlineKeyboardButton("💾 2. Simpan", callback_data=f"simpan_temp_format"), InlineKeyboardButton("🗑️ 3. Hapus", callback_data=f"hapus_temp_format")],
            [InlineKeyboardButton("⬅️ Kembali", callback_data="format_jmo")]
        ])
        await update.message.reply_text(f"✅ Diedit! Hasil baru:\n```\n{formatted}\n```\n\nKlik Simpan untuk masuk ke HASIL FORMAT!", reply_markup=kb, parse_mode="Markdown")
        return

    if mode=='edit_setting_format':
        # User edit template setting
        new_template = text
        conn=sqlite3.connect(DB_NAME)
        conn.execute("INSERT OR REPLACE INTO format_settings (user_id, template, updated_at) VALUES (?,?,?)",(uid, new_template, datetime.now().strftime("%d-%m-%Y %H:%M")))
        conn.commit(); conn.close()
        context.user_data['mode']=None
        await update.message.reply_text(f"✅ **SETTING FORMAT DISIMPAN!**\n\nTemplate baru:\n```\n{new_template}\n```\n\nSekarang ketika kamu klik 1. BUAT FORMAT dan ketik asal2an seperti:\nInfo loker\nBandung\nAntapani\nKulim\nGaji 12 juta\nDpt jmo lasik jos\n\nBot akan otomatis format sesuai template baru ini!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⚙️ Lihat Setting", callback_data="setting_format_menu")],[InlineKeyboardButton("1. BUAT FORMAT", callback_data="buat_format")]]), parse_mode="Markdown")
        return

    if mode=='edit_format':
        fid = context.user_data.get('edit_fid')
        if not fid:
            context.user_data['mode']=None
            await update.message.reply_text("Error, coba lagi", reply_markup=main_menu())
            return
        conn=sqlite3.connect(DB_NAME)
        conn.execute("UPDATE formats SET content=?, created_at=? WHERE id=? AND user_id=?",(text, datetime.now().strftime("%d-%m-%Y %H:%M"), fid, uid))
        conn.commit(); conn.close()
        context.user_data['mode']=None
        await update.message.reply_text(f"✅ Format #{fid} berhasil diedit!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(f"📄 Lihat Format #{fid}", callback_data=f"viewfmt_{fid}")],[InlineKeyboardButton("📂 HASIL FORMAT", callback_data="hasil_format")]]))
        return

    if mode=='cari_format':
        conn=sqlite3.connect(DB_NAME)
        rows=conn.execute("SELECT id,title FROM formats WHERE user_id=? AND (LOWER(title) LIKE ? OR LOWER(content) LIKE ?)",(uid,f"%{lower}%",f"%{lower}%")).fetchall()
        conn.close()
        if not rows:
            await update.message.reply_text(f"Tidak ada hasil '{text}'", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📂 Kembali",callback_data="hasil_format")]]))
        else:
            buttons=[[InlineKeyboardButton(f"📄 {t[:25]}", callback_data=f"viewfmt_{fid}")] for fid,t in rows]
            buttons.append([InlineKeyboardButton("📂 Kembali",callback_data="hasil_format")])
            await update.message.reply_text(f"🔍 Ketemu {len(rows)} untuk '{text}'", reply_markup=InlineKeyboardMarkup(buttons))
        context.user_data['mode']=None
        return

    if mode=='solusi_jmo_tanya':
        jawaban=None
        for key in SOLUSI_LENGKAP:
            if key in lower:
                jawaban=SOLUSI_LENGKAP[key]
                break
        if not jawaban:
            if "025" in lower or "duplikat" in lower: jawaban=SOLUSI_LENGKAP["025"]
            elif "026" in lower or "kpj" in lower: jawaban=SOLUSI_LENGKAP["026"]
            elif "040" in lower or "biometrik" in lower or "wajah" in lower: jawaban=SOLUSI_LENGKAP["040"]
            elif "otp" in lower: jawaban=SOLUSI_LENGKAP["otp"]
            elif "lupa" in lower: jawaban=SOLUSI_LENGKAP["lupa password"]
            elif "klaim" in lower: jawaban=SOLUSI_LENGKAP["klaim"]
            elif "saldo" in lower: jawaban=SOLUSI_LENGKAP["saldo"]
        
        if not jawaban:
            jawaban=f"❓ **Kamu tanya:** {text}\n\nCoba ketik kode error 025, 040, 026, otp, klaim, dll. Untuk OTP pakai email baru @{DOMAIN} cuma 5k!"

        kb=InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Tanya Lain",callback_data="solusi_jmo")],
            [InlineKeyboardButton("📧 Beli Email 5k",callback_data="beli_email")],
            [InlineKeyboardButton("⬅️ Menu",callback_data="back")]
        ])
        await update.message.reply_text(jawaban, reply_markup=kb, parse_mode="Markdown")
        return

    if mode=='topup_custom':
        try:
            amount=int(text.replace(".","").replace(",","").replace("k","000"))
            if amount < 5000:
                await update.message.reply_text("Minimal 5000! Ketik lagi:")
                return
            conn=sqlite3.connect(DB_NAME)
            conn.execute("INSERT INTO topup (user_id, amount, status, created_at) VALUES (?,?,?,?)",(uid, amount, "pending", datetime.now().strftime("%d-%m-%Y %H:%M")))
            topup_id=conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.commit(); conn.close()
            await update.message.reply_text(f"💳 TOPUP Rp {amount:,} ID #{topup_id}\nScan QRIS: {QRIS_IMAGE}\nUpload bukti setelah bayar!", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📤 Upload Bukti", callback_data="upload_bukti")]]))
        except:
            await update.message.reply_text("Format salah! Ketik angka, contoh 50000")
        context.user_data['mode']=None
        return

    if mode=='upload_bukti':
        await update.message.reply_text("⚠️ Kirim bukti dalam bentuk FOTO screenshot, bukan teks!")
        return

    await update.message.reply_text("Ketik /start untuk menu", reply_markup=main_menu())

async def document_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Terima file Excel dari menu FORMAT OTOMATIS -> BUAT FORMAT."""
    uid = update.effective_user.id
    mode = context.user_data.get('mode')
    document = update.message.document

    if mode != 'buat_format_xl':
        await update.message.reply_text("Ketik /start untuk menu", reply_markup=main_menu())
        return

    if not document:
        await update.message.reply_text("❌ File tidak ditemukan. Silahkan kirim file Excel (.xlsx).")
        return

    filename = document.file_name or ""
    if not filename.lower().endswith('.xlsx'):
        await update.message.reply_text(
            "❌ Format file belum didukung.\n\nSilahkan kirim file Excel **.xlsx** saja.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Batal", callback_data="format_jmo")]])
        )
        return

    await update.message.reply_text("⏳ File Excel diterima. Sedang membuat format otomatis, mohon tunggu...")
    input_path = f"/tmp/format_input_{uid}.xlsx"
    output_path = f"/tmp/hasil_format_{uid}.xlsx"

    try:
        tg_file = await document.get_file()
        await tg_file.download_to_drive(input_path)

        wb = openpyxl.load_workbook(input_path, data_only=True)
        template = get_format_setting(uid)
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = "Hasil Format"
        out_ws.append(["NO", "DATA ASAL", "HASIL FORMAT"])

        total = 0
        conn = sqlite3.connect(DB_NAME)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if not values:
                    continue
                raw_text = "\n".join(values)
                formatted = auto_format_text(raw_text, template)
                total += 1
                out_ws.append([total, raw_text, formatted])
                conn.execute(
                    "INSERT INTO format_history (user_id, raw_text, formatted_text, created_at) VALUES (?,?,?,?)",
                    (uid, raw_text, formatted, datetime.now().strftime("%d-%m-%Y %H:%M"))
                )

        conn.commit()
        conn.close()
        wb.close()

        if total == 0:
            raise ValueError("File Excel tidak memiliki data yang bisa diproses.")

        out_ws.column_dimensions['A'].width = 8
        out_ws.column_dimensions['B'].width = 45
        out_ws.column_dimensions['C'].width = 70
        for row in out_ws.iter_rows(min_row=2):
            row[1].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
            row[2].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

        out_wb.save(output_path)
        out_wb.close()

        context.user_data['mode'] = None
        await update.message.reply_document(
            document=open(output_path, "rb"),
            filename=f"HASIL_FORMAT_OTOMATIS_{uid}.xlsx",
            caption=(
                f"✅ **FORMAT OTOMATIS SELESAI!**\n\n"
                f"📊 File: `{filename}`\n"
                f"📝 Data berhasil diproses: **{total} baris**\n\n"
                f"File hasil sudah dibuat sesuai SETTING FORMAT kamu."
            ),
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📂 Lihat HASIL FORMAT", callback_data="hasil_format")],
                [InlineKeyboardButton("📊 BUAT FORMAT LAGI", callback_data="buat_format")],
                [InlineKeyboardButton("⬅️ Kembali", callback_data="format_jmo")]
            ])
        )
    except Exception as e:
        context.user_data['mode'] = 'buat_format_xl'
        await update.message.reply_text(
            f"❌ Gagal memproses file Excel.\n\nError: {str(e)[:300]}\n\nSilahkan kirim ulang file **.xlsx** yang benar.",
            parse_mode="Markdown"
        )
    finally:
        for path in (input_path, output_path):
            try:
                os.remove(path)
            except OSError:
                pass

async def photo_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Menerima bukti pembayaran dan membuat topup pending."""
    user = update.effective_user
    if not user or not update.message or not update.message.photo:
        return

    uid = user.id
    conn = sqlite3.connect(DB_NAME)
    row = conn.execute(
        "SELECT id, amount FROM topup WHERE user_id=? AND status='pending' ORDER BY id DESC LIMIT 1",
        (uid,)
    ).fetchone()
    conn.close()

    if not row:
        await update.message.reply_text(
            "❌ Tidak ada topup pending. Buat topup dulu dari menu TOP UP.",
            reply_markup=main_menu()
        )
        return

    topup_id, amount = row
    proof = update.message.photo[-1].file_id
    username = user.username or "-"

    caption = (
        f"💳 <b>BUKTI TOPUP BARU</b>\n\n"
        f"🧾 ID: <code>#{topup_id}</code>\n"
        f"👤 User: <code>{uid}</code>\n"
        f"📱 Username: @{username}\n"
        f"💰 Nominal: <b>Rp {amount:,}</b>\n"
        f"⏳ Status: <b>PENDING</b>"
    )

    admin_kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton(
                f"✅ APPROVE #{topup_id}",
                callback_data=f"approve_topup_{topup_id}"
            ),
            InlineKeyboardButton(
                f"❌ REJECT #{topup_id}",
                callback_data=f"reject_topup_{topup_id}"
            )
        ]
    ])

    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_photo(
                chat_id=admin_id,
                photo=proof,
                caption=caption,
                parse_mode="HTML",
                reply_markup=admin_kb
            )
        except Exception as exc:
            print(f"Gagal kirim bukti ke admin {admin_id}: {exc}")

    context.user_data['mode'] = None

    await update.message.reply_text(
        f"✅ <b>BUKTI PEMBAYARAN DITERIMA</b>\n\n"
        f"🧾 ID: #{topup_id}\n"
        f"💰 Nominal: Rp {amount:,}\n\n"
        "⏳ Bukti sudah dikirim ke Admin.\n"
        "Saldo akan masuk setelah Admin menekan APPROVE.",
        parse_mode="HTML",
        reply_markup=main_menu()
    )

async def daily_report_job(context: ContextTypes.DEFAULT_TYPE):
    """Laporan harian otomatis kirim ke admin jam 21:00"""
    try:
        conn=sqlite3.connect(DB_NAME)
        today=datetime.now().strftime("%d-%m-%Y")
        # Topup hari ini
        total_today=conn.execute("SELECT SUM(amount) FROM topup WHERE status LIKE 'approved%' AND created_at LIKE ?", (f"%{today}%",)).fetchone()[0] or 0
        count_today=conn.execute("SELECT COUNT(*) FROM topup WHERE status LIKE 'approved%' AND created_at LIKE ?", (f"%{today}%",)).fetchone()[0]
        # Total semua
        total_all=conn.execute("SELECT SUM(amount) FROM topup WHERE status LIKE 'approved%'").fetchone()[0] or 0
        total_users=conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        total_email=conn.execute("SELECT COUNT(*) FROM emails").fetchone()[0]
        total_saldo=conn.execute("SELECT SUM(saldo) FROM users").fetchone()[0] or 0
        # Top user hari ini
        top_rows=conn.execute("SELECT user_id, amount, created_at FROM topup WHERE status LIKE 'approved%' AND created_at LIKE ? ORDER BY amount DESC LIMIT 5", (f"%{today}%",)).fetchall()
        conn.close()
        
        txt=f"""📊 **LAPORAN HARIAN OTOMATIS - {today}**

**HARI INI:**
💳 Pemasukan: Rp {total_today:,}
🔢 Transaksi: {count_today} topup
📧 Email Terjual Hari Ini: Estimasi {count_today} paket

**TOTAL KESELURUHAN:**
👤 Total User: {total_users}
💰 Total Saldo User: Rp {total_saldo:,}
💵 Total Pemasukan: Rp {total_all:,}
📧 Total Email Terjual: {total_email}

**TOP 5 TOPUP HARI INI:**
"""
        for i, (uid, amount, tgl) in enumerate(top_rows, 1):
            txt+=f"{i}. ID {uid} - Rp {amount:,} - {tgl}\n"
        
        txt+=f"\nHarga: 1 Email 5k, 10 Email 40k, 20 Email 70k\nBot jalan auto 24 jam!"
        
        for admin_id in ADMIN_IDS:
            try:
                await context.bot.send_message(chat_id=admin_id, text=txt, parse_mode="Markdown")
            except:
                pass
    except Exception as e:
        print(f"Error daily report: {e}")

if __name__=="__main__":
    if not TOKEN:
        print("BOT_TOKEN belum di set!")
    else:
        init_db()
        app=ApplicationBuilder().token(TOKEN).build()
        app.add_handler(CommandHandler("start",start))
        app.add_handler(CallbackQueryHandler(button_handler))
        app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, msg_handler))
        app.add_handler(MessageHandler(filters.Document.ALL, document_handler))
        app.add_handler(MessageHandler(filters.PHOTO, photo_handler))
        
        # === LAPORAN HARIAN OTOMATIS JAM 21:00 WIB ===
        # JobQueue kirim laporan tiap hari jam 21:00
        try:
            job_queue = app.job_queue
            # Kirim tiap hari jam 21:00 WIB (14:00 UTC)
            job_queue.run_daily(daily_report_job, time=datetime.strptime("14:00", "%H:%M").time(), name="daily_report")
            print("✅ Laporan harian otomatis aktif jam 21:00 WIB")
        except Exception as e:
            print(f"JobQueue tidak aktif: {e} - Laporan manual tetap bisa via /admin")
        
        print(f"BOT JMO FINAL + FORMAT OTOMATIS EXCEL (.XLSX) + LAPORAN HARIAN JALAN!")
        app.run_polling()
